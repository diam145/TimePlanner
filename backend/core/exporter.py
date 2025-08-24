# backend/core/exporter.py
"""
Professional Excel Export Module for TimePlanner
Creates schedules in the format matching modelExcel.xlsx
"""

import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
import calendar

def generate_professional_schedule_export(start_date: datetime, end_date: datetime, schedule_data: dict, filename: str = None):
    """
    Generate a professional Excel schedule export matching the modelExcel.xlsx format
    
    Args:
        start_date (datetime): Start date of the schedule period
        end_date (datetime): End date of the schedule period  
        schedule_data (dict): Dictionary with employee schedules
        filename (str): Optional custom filename
    
    Returns:
        str: Path to the created Excel file
    """
    
    # Calculate months to display
    months_to_display = []
    current = start_date.replace(day=1)
    while current <= end_date:
        months_to_display.append(current)
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)
    
    # Limit to 2 months for the dual-month format
    if len(months_to_display) > 2:
        months_to_display = months_to_display[:2]
    
    # Generate filename if not provided
    if not filename:
        if len(months_to_display) == 2:
            month1_name = get_french_month_name(months_to_display[0].month)
            month2_name = get_french_month_name(months_to_display[1].month)
            year = months_to_display[0].year
            # Add timestamp to avoid file conflicts
            timestamp = datetime.now().strftime("%H%M%S")
            filename = f"Horaire_garde_sages_femmes_{month1_name}_{month2_name}_{year}_{timestamp}.xlsx"
        else:
            month_name = get_french_month_name(months_to_display[0].month)
            year = months_to_display[0].year
            timestamp = datetime.now().strftime("%H%M%S")
            filename = f"Horaire_garde_sages_femmes_{month_name}_{year}_{timestamp}.xlsx"
    
    # Create archive directory path
    archive_dir = os.path.join(os.path.dirname(__file__), 'data', 'archive')
    os.makedirs(archive_dir, exist_ok=True)
    file_path = os.path.join(archive_dir, filename)
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Horaires de Garde"
    
    # Set page orientation and margins
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    
    # Create the professional format
    create_header_section(ws)
    create_schedule_grid(ws, months_to_display, schedule_data)
    create_summary_section(ws, schedule_data, months_to_display)
    
    # Apply styling
    apply_professional_styling(ws, months_to_display)
    
    # Save the file
    wb.save(file_path)
    print(f"✅ Professional schedule export saved: {file_path}")
    
    return file_path

def create_header_section(ws):
    """Create the professional header section with organization details and legend"""
    
    # Main Title - Large and merged across full width
    ws['A1'] = "Horaire de garde des sages femmes"
    ws.merge_cells('A1:AL1')  # Merge across full width
    
    # Year in top right
    ws['AM1'] = "2025"
    
    # Organization details section (Left side)
    ws['A3'] = "Maison de naissance de la Rivière"
    ws['A4'] = "1275, rue Saint-Jean-Baptiste, Nicolet (Québec) J3T 1V4"
    ws['A5'] = "Tél.: 819 293-2071, poste 56221"
    ws['A6'] = "Sans frais (1 800 263-2572, poste 56221)"
    
    # Legend section (Right side with cyan background)
    ws['X3'] = "☐ Garde"
    ws['X4'] = "☐ = Soutien de jour seulement"
    ws['X5'] = "☐ Congé"
    ws['Y3'] = "RP = Rencontre Prénatale"
    ws['Y4'] = "E 9h ou E 17h: la garde Fini à 9h ou à 17h"
    ws['Y5'] = "F 9h ou F 17h: la garde Fin à 9h ou à 17h"
    
    # Create merged cells for organization info
    ws.merge_cells('A3:W3')  # Organization name
    ws.merge_cells('A4:W4')  # Address
    ws.merge_cells('A5:W5')  # Phone
    ws.merge_cells('A6:W6')  # Toll-free
    
    # Create merged cells for legend
    ws.merge_cells('X3:AL6')  # Legend area
    
    return 7  # Return the row where the schedule grid should start

def create_schedule_grid(ws, months_to_display, schedule_data):
    """Create the main schedule grid with dual-month layout starting from row 7, column A empty"""
    
    start_row = 7  # Start from row 7 as requested
    current_row = start_row
    
    # Create month headers (starting from column B, leaving A empty)
    if len(months_to_display) == 2:
        # Dual month layout
        month1 = months_to_display[0]
        month2 = months_to_display[1]
        
        # Month 1 header (start from column B)
        month1_name = get_french_month_name(month1.month).upper()
        ws.cell(row=current_row, column=2, value=f"{month1_name}")  # Column B
        
        # Calculate days in month 1
        days_in_month1 = calendar.monthrange(month1.year, month1.month)[1]
        month1_end_col = 1 + days_in_month1  # Account for empty column A
        ws.merge_cells(f'B{current_row}:{get_column_letter(month1_end_col)}{current_row}')
        
        # Month 2 header
        month2_name = get_french_month_name(month2.month).upper()
        month2_start_col = month1_end_col + 1
        ws.cell(row=current_row, column=month2_start_col, value=f"{month2_name}")
        
        # Calculate days in month 2
        days_in_month2 = calendar.monthrange(month2.year, month2.month)[1]
        month2_end_col = month2_start_col + days_in_month2 - 1
        ws.merge_cells(f'{get_column_letter(month2_start_col)}{current_row}:{get_column_letter(month2_end_col)}{current_row}')
        
        current_row += 1
        
        # Day numbers row (starting from column B, leaving A empty)
        col_index = 2  # Start from column B
        
        # Month 1 days
        for day in range(1, days_in_month1 + 1):
            ws.cell(row=current_row, column=col_index, value=day)
            col_index += 1
        
        # Month 2 days
        for day in range(1, days_in_month2 + 1):
            ws.cell(row=current_row, column=col_index, value=day)
            col_index += 1
        
        current_row += 1
        
        # Day names row (starting from column B, leaving A empty)
        col_index = 2  # Start from column B
        
        # Month 1 day names
        for day in range(1, days_in_month1 + 1):
            date_obj = datetime(month1.year, month1.month, day)
            day_name = get_french_day_abbrev(date_obj.weekday())
            ws.cell(row=current_row, column=col_index, value=day_name)
            col_index += 1
        
        # Month 2 day names
        for day in range(1, days_in_month2 + 1):
            date_obj = datetime(month2.year, month2.month, day)
            day_name = get_french_day_abbrev(date_obj.weekday())
            ws.cell(row=current_row, column=col_index, value=day_name)
            col_index += 1
        
        current_row += 1
        
        # Employee header (in column A as usual)
        ws.cell(row=current_row, column=1, value="Sages femmes")
        current_row += 1
        
        # Employee rows (employee names in column A, schedule data starting from column B)
        for employee_name, employee_data in schedule_data.items():
            # Employee name in column A
            ws.cell(row=current_row, column=1, value=employee_name)
            
            col_index = 2  # Start schedule data from column B
            
            # Month 1 schedule
            for day in range(1, days_in_month1 + 1):
                date_key = f"{month1.year}-{month1.month:02d}-{day:02d}"
                schedule_value = get_schedule_value_for_date(employee_data, date_key)
                ws.cell(row=current_row, column=col_index, value=schedule_value)
                col_index += 1
            
            # Month 2 schedule
            for day in range(1, days_in_month2 + 1):
                date_key = f"{month2.year}-{month2.month:02d}-{day:02d}"
                schedule_value = get_schedule_value_for_date(employee_data, date_key)
                ws.cell(row=current_row, column=col_index, value=schedule_value)
                col_index += 1
            
            current_row += 1
    
    return current_row

def create_summary_section(ws, schedule_data, months_to_display):
    """Create the summary section with totals"""
    
    # This will be implemented to show totals per employee
    # For now, we'll add a simple summary
    pass

def apply_professional_styling(ws, months_to_display):
    """Apply sophisticated styling to match the professional Excel model"""
    
    # Professional Fonts matching the model
    title_font_large = Font(name='Franklin Gothic Medium', size=36, bold=True)  # Large title
    section_header_font = Font(name='Berlin Sans FB Demi', size=14, bold=True, color='FFFFFF')  # White text on dark
    month_header_font = Font(name='Berlin Sans FB Demi', size=12, bold=True)
    normal_font = Font(name='Franklin Gothic Book', size=10)
    small_font = Font(name='Franklin Gothic Book', size=9)
    contact_font = Font(name='Arial', size=9)
    
    # Alignments
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    top_left_align = Alignment(horizontal='left', vertical='top')
    
    # Professional Colors matching the model
    cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')  # Bright cyan
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')  # White
    dark_gray_fill = PatternFill(start_color='404040', end_color='404040', fill_type='solid')  # Dark gray headers
    light_blue_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')  # Light blue
    
    # Professional Borders
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    no_border = Border()
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Apply main title styling (Row 1)
    title_cell = ws['A1']
    title_cell.font = title_font_large
    title_cell.alignment = center_align
    title_cell.border = no_border
    
    # Apply organization info styling (Rows 3-6)
    for row in range(3, 7):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = contact_font
            cell.alignment = top_left_align
            cell.border = no_border
    
    # Apply legend styling with cyan highlight (Right side)
    legend_columns = ['X', 'Y', 'Z', 'AA', 'AB']
    for row in range(3, 8):
        for col_letter in legend_columns:
            try:
                cell = ws[f'{col_letter}{row}']
                cell.font = small_font
                cell.alignment = left_align
                cell.fill = cyan_fill
                cell.border = thin_border
            except:
                pass
    
    # Find and style month headers with dark background
    month_row = 7  # Month headers now start at row 7
    for col in range(1, max_col + 1):
        cell = ws.cell(row=month_row, column=col)
        if cell.value and str(cell.value).upper() in ['SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DÉCEMBRE']:
            cell.font = month_header_font
            cell.alignment = center_align
            cell.fill = dark_gray_fill
            cell.border = thick_border
    
    # Style day number headers
    day_numbers_row = 8  # Day numbers at row 8
    for col in range(1, max_col + 1):
        cell = ws.cell(row=day_numbers_row, column=col)
        if cell.value and str(cell.value).isdigit():
            cell.font = normal_font
            cell.alignment = center_align
            cell.fill = dark_gray_fill
            cell.border = thin_border
    
    # Style day names headers (L, Ma, Me, J, V, S, D)
    day_names_row = 9  # Day names at row 9
    for col in range(1, max_col + 1):
        cell = ws.cell(row=day_names_row, column=col)
        cell.font = normal_font
        cell.alignment = center_align
        cell.fill = dark_gray_fill
        cell.border = thin_border
    
    # Style "Sages femmes" section header
    employee_header_row = 10  # Employee header at row 10
    for col in range(1, max_col + 1):
        cell = ws.cell(row=employee_header_row, column=col)
        if col == 1 and cell.value:
            cell.font = section_header_font
            cell.alignment = left_align
            cell.fill = dark_gray_fill
            cell.border = thick_border
    
    # Apply alternating cyan/white pattern to employee schedule grid
    employee_start_row = 11  # Employee data starts at row 11
    for row in range(employee_start_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.border = thin_border
            
            # Determine alternating pattern based on day column
            if col == 1:  # Employee name column
                cell.alignment = left_align
                cell.fill = white_fill
            else:
                cell.alignment = center_align
                
                # Create alternating cyan/white pattern for days
                # Every other day gets cyan background
                day_index = (col - 2) % 7  # Cycle through week
                if day_index % 2 == 0:  # Even days get cyan
                    cell.fill = cyan_fill
                else:  # Odd days stay white
                    cell.fill = white_fill
    
    # Set professional column widths
    ws.column_dimensions['A'].width = 15  # Employee names column
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4.5  # Day columns
    
    # Set professional row heights
    ws.row_dimensions[1].height = 45  # Title row
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 22  # Standard row height
    
    # Add thick borders around major sections
    # Title section border
    for col in range(1, max_col + 1):
        ws.cell(row=1, column=col).border = Border(
            top=Side(style='thick', color='000000'),
            bottom=Side(style='thick', color='000000')
        )
    
    # Schedule grid outline (starting from row 7)
    for row in range(7, max_row + 1):
        # Left border
        ws.cell(row=row, column=1).border = Border(
            left=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )
        # Right border
        ws.cell(row=row, column=max_col).border = Border(
            right=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000')
        )
    
    # Bottom border for schedule grid
    for col in range(1, max_col + 1):
        ws.cell(row=max_row, column=col).border = Border(
            bottom=Side(style='thick', color='000000'),
            top=Side(style='thin', color='000000'),
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000')
        )
    
    # Alignments
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    top_align = Alignment(horizontal='center', vertical='top')
    
    # Professional Colors matching the model
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
    black_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    light_blue_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')
    
    # Professional Borders (743 cells have borders in the model)
    grid_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    no_border = Border()
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 1. Apply large title styling (Row 1) - Franklin Gothic Medium 36pt
    title_cell = ws['A1']
    title_cell.font = title_font_large
    title_cell.alignment = center_align
    title_cell.fill = white_fill
    title_cell.border = no_border
    
    # 2. Year styling (merge cells for title area)
    year_cell = ws.cell(row=1, column=max_col-5)  # Position year at the right
    if year_cell.value:
        year_cell.font = title_font_large
        year_cell.alignment = center_align
        year_cell.fill = white_fill
    
    # 3. Organization header styling (Rows 3-6) - Berlin Sans FB Demi
    for row in range(3, 7):
        for col in range(1, 6):  # Organization info area
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = contact_font  # Use contact_font instead of undefined header_font
                cell.alignment = left_align
                cell.fill = white_fill
                cell.border = no_border
    
    # 4. Legend styling (right side) - Cyan background, Franklin Gothic Book
    for row in range(3, 8):
        for col_letter in ['X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD']:
            try:
                cell = ws[f'{col_letter}{row}']
                if cell.value:
                    cell.font = small_font
                    cell.alignment = left_align
                    cell.fill = cyan_fill
                    cell.border = grid_border
            except:
                pass
    
    # 5. Month headers styling - Berlin Sans FB Demi, bold
    month_row = 8
    for col in range(2, max_col + 1):
        cell = ws.cell(row=month_row, column=col)
        if cell.value and any(month in str(cell.value).upper() for month in ['SEPTEMBRE', 'OCTOBRE', 'NOVEMBRE', 'DÉCEMBRE']):
            cell.font = Font(name='Berlin Sans FB Demi', size=12, bold=True, color='FFFFFF')  # White text on black
            cell.alignment = center_align
            cell.fill = black_fill
            cell.border = grid_border
    
    # 6. Day numbers row styling - Franklin Gothic Book
    day_row = 9
    for col in range(2, max_col + 1):
        cell = ws.cell(row=day_row, column=col)
        if cell.value and str(cell.value).isdigit():
            cell.font = normal_font
            cell.alignment = center_align
            cell.fill = white_fill
            cell.border = grid_border
    
    # 7. Day names row styling (L, Ma, Me, etc.) - Franklin Gothic Book
    day_names_row = 10
    for col in range(2, max_col + 1):
        cell = ws.cell(row=day_names_row, column=col)
        if cell.value:
            cell.font = small_font
            cell.alignment = center_align
            cell.fill = white_fill
            cell.border = grid_border
    
    # 8. Employee section header styling - Berlin Sans FB Demi
    employee_header_row = 11
    header_cell = ws.cell(row=employee_header_row, column=1)
    if header_cell.value:
        header_cell.font = Font(name='Berlin Sans FB Demi', size=12, bold=True, color='FFFFFF')
        header_cell.alignment = left_align
        header_cell.fill = black_fill
        header_cell.border = grid_border
    
    # 9. Employee data grid with alternating colors and borders
    employee_start_row = 12
    for row in range(employee_start_row, max_row + 1):
        employee_row_index = row - employee_start_row
        is_even_row = employee_row_index % 2 == 0
        
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = grid_border  # All data cells have borders
            
            if col == 1:  # Employee name column
                cell.font = normal_font
                cell.alignment = left_align
                # Alternating background colors
                if is_even_row:
                    cell.fill = light_blue_fill
                else:
                    cell.fill = white_fill
            else:  # Schedule data columns
                cell.font = normal_font
                cell.alignment = center_align
                # Alternating background colors
                if is_even_row:
                    cell.fill = light_blue_fill
                else:
                    cell.fill = white_fill
                
                # Special styling for specific values
                if cell.value == 'X':
                    cell.font = Font(name='Franklin Gothic Book', size=10, bold=True)
                elif cell.value == 'S':
                    cell.font = Font(name='Franklin Gothic Book', size=10, bold=True)
    
    # 10. Summary section at bottom with borders
    summary_start_row = max_row - 5  # Assuming last 5 rows are summary
    for row in range(summary_start_row, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.font = small_font
                cell.alignment = center_align
                cell.fill = white_fill
                cell.border = grid_border
    
    # 11. Set professional column widths
    ws.column_dimensions['A'].width = 15  # Employee names
    for col in range(2, max_col + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 3.5  # Compact day columns
    
    # 12. Set professional row heights
    ws.row_dimensions[1].height = 45  # Large title row
    for row in range(2, max_row + 1):
        ws.row_dimensions[row].height = 18  # Standard row height
    
    # 13. Merge cells for headers (45 merged ranges in the model)
    # Title area merge
    try:
        ws.merge_cells('A1:F1')  # Title merge
        # Month header merges would be calculated based on days in month
        # This would require more sophisticated logic based on actual data
    except:
        pass  # Skip if merge conflicts
    
    # Apply day number row styling
    day_row = 9  # Day numbers row
    for col in range(1, max_col + 1):
        cell = ws.cell(row=day_row, column=col)
        if cell.value and str(cell.value).isdigit():
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = no_border
            
            # Check if it's a weekend (Saturday/Sunday)
            # This would need date logic to determine weekends
            # For now, apply weekend styling to specific columns if needed

    # Apply employee name column styling
    for row in range(10, max_row + 1):
        name_cell = ws.cell(row=row, column=1)  # Column A for names
        if name_cell.value:
            name_cell.font = normal_font
            name_cell.alignment = left_align
            name_cell.border = no_border

    # Apply schedule grid styling
    for row in range(10, max_row + 1):
        for col in range(2, max_col + 1):  # Skip the name column
            cell = ws.cell(row=row, column=col)
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = no_border            # Apply weekend highlighting if needed
            # This could be enhanced with actual date checking
    
    # Set column widths to match the template
    ws.column_dimensions['A'].width = 15  # Employee names
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4  # Day columns
    
    # Set row heights
    for row in range(1, max_row + 1):
        ws.row_dimensions[row].height = 20
        if row > 10:  # Data rows
            # Check if this is a weekend column
            # This would need to be calculated based on the date
            pass
    
    # Set column widths
    ws.column_dimensions['A'].width = 20  # Employee names
    for col in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 4  # Day columns

def get_schedule_value_for_date(employee_data, date_key):
    """Get the schedule value for an employee on a specific date"""
    
    # Check if there's schedule data for this date
    if 'schedules' in employee_data and date_key in employee_data['schedules']:
        schedule_info = employee_data['schedules'][date_key]
        if isinstance(schedule_info, dict):
            activity_type = schedule_info.get('activity_type', '')
            # Format according to the display logic (add time for D and F)
            if activity_type == 'D':
                start_time = schedule_info.get('start_time', '17:00')
                if start_time == '17:00':
                    return 'D\n17h'
                else:
                    return 'D\n16h'
            elif activity_type == 'F':
                end_time = schedule_info.get('end_time', '09:00')
                if end_time == '09:00':
                    return 'F\n9h'
                else:
                    return 'F\n8h'
            else:
                return activity_type
        else:
            return str(schedule_info)
    
    # Apply F\n9h injection logic (same as frontend)
    # Check if previous day was a night shift
    try:
        from datetime import datetime, timedelta
        current_date = datetime.strptime(date_key, '%Y-%m-%d')
        prev_date = current_date - timedelta(days=1)
        prev_date_key = prev_date.strftime('%Y-%m-%d')
        
        if ('schedules' in employee_data and 
            prev_date_key in employee_data['schedules']):
            prev_schedule = employee_data['schedules'][prev_date_key]
            if isinstance(prev_schedule, dict) and prev_schedule.get('is_night_shift', False):
                # Inject F\n9h for end of night shift
                return 'F\n9h'
    except (ValueError, KeyError):
        pass
    
    return ''

def get_french_month_name(month_num):
    """Get French month name"""
    french_months = {
        1: 'Janvier', 2: 'Février', 3: 'Mars', 4: 'Avril',
        5: 'Mai', 6: 'Juin', 7: 'Juillet', 8: 'Août',
        9: 'Septembre', 10: 'Octobre', 11: 'Novembre', 12: 'Décembre'
    }
    return french_months.get(month_num, 'Inconnu')

def get_french_day_abbrev(weekday):
    """Get French day abbreviation (Monday=0)"""
    french_days = ['L', 'Ma', 'Me', 'J', 'V', 'S', 'D']  # Monday to Sunday
    return french_days[weekday]

# Legacy function for backward compatibility
def export_weekly_report(df: pd.DataFrame, week_start: datetime):
    """Legacy export function - now calls the professional export"""
    
    # Convert DataFrame to schedule_data format
    schedule_data = {}
    for _, row in df.iterrows():
        employee_name = row.get('Name', row.get('name', 'Unknown'))
        schedule_data[employee_name] = {
            'schedules': {}
        }
        
        # Convert weekly data to date-based format
        days = ['Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday']
        for i, day in enumerate(days):
            if day in row and row[day]:
                date_obj = week_start + timedelta(days=i)
                date_key = date_obj.strftime('%Y-%m-%d')
                schedule_data[employee_name]['schedules'][date_key] = row[day]
    
    end_date = week_start + timedelta(days=6)
    
    return generate_professional_schedule_export(
        start_date=week_start,
        end_date=end_date,
        schedule_data=schedule_data
    )
