import sqlite3
import pandas as pd
from datetime import datetime, timedelta
from models import Session, Worker


def get_start_of_week(day:datetime):
    days_since_sunday = (day.weekday() + 1) % 7  # Sunday
    week_start = day - timedelta(days=days_since_sunday)

    # Optional: reset time to 00:00
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    return week_start


def get_last_sunday():
    today = datetime.today()
    days_since_sunday = (today.weekday() + 1) % 7  # Sunday
    last_sunday = today - timedelta(days=days_since_sunday)

    # Optional: reset time to 00:00
    last_sunday = last_sunday.replace(hour=0, minute=0, second=0, microsecond=0)

    return last_sunday


def load_all_sessions():
    conn = sqlite3.connect('./data/workers.db')
    df = pd.read_sql_query("SELECT * FROM sessions", conn)
    conn.close()
    return df

def load_sessions_with_names():
    conn = sqlite3.connect('./data/workers.db')
    df = pd.read_sql_query("""
        SELECT s.worker_id, w.name, s.start, s.end, s.activity
        FROM sessions s
        JOIN workers w ON s.worker_id = w.id
        ORDER BY w.name ASC, s.start ASC
    """, conn)
    conn.close()
    return df

def export_weekly_report(df: pd.DataFrame, week_start: datetime):
    week_start = get_start_of_week(week_start)
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    # Step 1: Build filename
    filename = f"weekly_report_{week_start.strftime('%Y-%m-%d')}.xlsx"
    filepath = f"./data/archive/{filename}"

    # Step 2: Export to Excel
    df.to_excel(filepath, index=False)

    # Step 3: Load Excel and apply text wrapping
    wb = load_workbook(filepath)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(filepath)
    print(f"Weekly report saved to: {filepath}")


def generate_weekly_table(df_workers, week_start: datetime) -> pd.DataFrame:
    # make sure we have the correct week start
    week_start = get_start_of_week(week_start)
    week_end = week_start + timedelta(days=6)

    columns = ["Name", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "TOTAL"]
    table_rows = []

    # Connect to database and get schedules for this week
    conn = sqlite3.connect('./data/workers.db')
    
    # Create the table if it doesn't exist
    conn.execute('''
        CREATE TABLE IF NOT EXISTS employee_schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            employee_id INTEGER,
            date DATE,
            day_of_week INTEGER,
            start_time TEXT,
            end_time TEXT,
            schedule_type TEXT DEFAULT 'work',
            activity_type TEXT DEFAULT 'X',
            is_night_shift BOOLEAN DEFAULT 0,
            shift_number INTEGER DEFAULT 1,
            hours_worked REAL DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (employee_id) REFERENCES workers (id)
        )
    ''')

    # Collect rows with worker name for sorting
    sortable_rows = []

    for worker_id in df_workers.index:
        name = df_workers.loc[worker_id]['name']
        
        # Get schedules for this worker for the week
        schedules = conn.execute('''
            SELECT date, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_number
            FROM employee_schedules
            WHERE employee_id = ? AND date >= ? AND date <= ? AND schedule_type = 'work'
            ORDER BY date, shift_number
        ''', (worker_id, week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d'))).fetchall()
        
        # Initialize row for this worker
        row = {day: "" for day in columns}
        row["Name"] = name
        
        # Track unique days worked (for proper night shift counting)
        days_worked = set()
        
        # Group schedules by date to handle multiple shifts per day
        schedules_by_date = {}
        for schedule in schedules:
            date_str = schedule[0]  # date column
            if date_str not in schedules_by_date:
                schedules_by_date[date_str] = []
            schedules_by_date[date_str].append(schedule)
        
        # Process each date
        for date_str, day_schedules in schedules_by_date.items():
            schedule_date = datetime.strptime(date_str, '%Y-%m-%d')
            day_name = schedule_date.strftime('%A')
            
            # For multiple shifts, show activity types separated by commas
            activity_types = []
            night_shift_end = None
            
            for schedule in day_schedules:
                _, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_number = schedule
                
                if activity_type:
                    activity_types.append(activity_type)
                    # Count this day as worked
                    days_worked.add(schedule_date.date())
                    
                    # For night shifts, also count the next day (but only if it's within this week)
                    if is_night_shift:
                        next_date = schedule_date + timedelta(days=1)
                        # Only count the next day if it falls within the current week
                        if week_start.date() <= next_date.date() <= week_end.date():
                            days_worked.add(next_date.date())
                
                # Handle night shift end marking for next day
                if is_night_shift and end_time:
                    next_date = schedule_date + timedelta(days=1)
                    if week_start <= next_date <= week_end:
                        next_day_name = next_date.strftime('%A')
                        if row[next_day_name] == "":
                            # Apply proper F formatting logic
                            end_hour = int(end_time.split(':')[0])
                            if end_hour < 12:  # Before noon, use F format
                                row[next_day_name] = f"F{end_time}"
                            else:  # At or after noon, use X format
                                row[next_day_name] = "X"
            
            # Set the activity for this day
            if activity_types:
                if len(activity_types) == 1:
                    row[day_name] = activity_types[0]
                else:
                    # Multiple shifts - show count or combined activities
                    row[day_name] = f"{len(activity_types)}×"  # or "/".join(activity_types)
        
        # Total is the number of unique days worked (including night shift spanning days)
        row["TOTAL"] = len(days_worked)
        sortable_rows.append(row)

    conn.close()

    # ✅ Sort rows by worker name
    sortable_rows.sort(key=lambda r: r["Name"])

    # Build final DataFrame
    df = pd.DataFrame(sortable_rows, columns=columns)
    return df
