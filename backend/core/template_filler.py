#!/usr/bin/env python3
"""
Template Filler for Excel Schedule Templates
Fills Excel templates with team member schedule data while preserving formatting.
"""

import openpyxl
from openpyxl.styles import Font, Alignment
import os
from datetime import datetime
from typing import Dict, List, Optional, Union


class ScheduleTemplateFiller:
    """
    A class to fill Excel schedule templates with team member data
    while preserving all existing formatting.
    """
    
    def __init__(self, template_path: str):
        """
        Initialize the template filler.
        
        Args:
            template_path: Path to the Excel template file
        """
        self.template_path = template_path
        self.workbook = None
        self.worksheet = None
        
    def load_template(self):
        """Load the Excel template."""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
        try:
            self.workbook = openpyxl.load_workbook(self.template_path)
            self.worksheet = self.workbook.active
            print(f"✅ Template loaded: {self.template_path}")
        except Exception as e:
            print(f"❌ Error loading template: {e}")
            raise
    
    def _detect_template_type(self) -> str:
        """
        Detect whether this is a schedule template or contact template.
        
        Returns:
            "schedule" or "contact"
        """
        # Check for schedule indicators
        for row in range(1, min(20, self.worksheet.max_row + 1)):
            for col in range(1, min(20, self.worksheet.max_column + 1)):
                cell_value = self.worksheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    cell_lower = cell_value.lower()
                    # Look for schedule template indicators
                    if any(term in cell_lower for term in ['horaire de garde', 'garde des sages', 'sages femmes']):
                        return "schedule"
                    # Look for contact-related terms (but only if no schedule terms found)
                    elif any(term in cell_lower for term in ['tél', 'fax', 'phone', 'maison']):
                        return "contact"
        
        # Default to schedule if unclear
        return "schedule"
    
    def _transform_schedule_symbol(self, symbol):
        """Transform database symbols to display symbols"""
        if symbol == 'V':
            return '-'  # Holiday symbol
        elif symbol == 'M':
            return 'M'  # Sickness symbol (Maladie)
        elif symbol in ['Maladie']:
            return 'M'  # Handle text "Maladie" as sickness
        else:
            return symbol  # Keep other symbols as is
    
    def _is_weekend_column(self, col_index, date_row=9):
        """Check if a column represents a weekend day"""
        try:
            # Get the date from row 9 (day numbers)
            day_value = self.worksheet.cell(row=date_row, column=col_index).value
            if day_value and isinstance(day_value, (int, float)):
                # For simplicity, assume we can determine weekends from day symbols in row 10
                day_symbol_cell = self.worksheet.cell(row=10, column=col_index)
                if day_symbol_cell.value:
                    day_symbol = str(day_symbol_cell.value).upper()
                    # Check if it's Saturday (S) or Sunday (D for Dimanche)
                    return day_symbol in ['S', 'D', 'SA', 'DI', 'SAM', 'DIM']
            return False
        except Exception as e:
            print(f"     ⚠️  Could not check weekend for column {col_index}: {e}")
            return False
    
    def safe_set_cell(self, row, col, value, font=None, alignment=None):
        """Safely set cell value (handles merged cells)"""
        try:
            cell = self.worksheet.cell(row=row, column=col)
            # Check if cell is part of a merged range
            if hasattr(cell, 'coordinate'):
                for merged_range in self.worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Unmerge the range first
                        self.worksheet.unmerge_cells(str(merged_range))
                        break
            
            # Now safely set the value
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = value
            
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
                
            return True
        except Exception as e:
            print(f"     ⚠️  Could not set cell ({row}, {col}): {e}")
            return False
    
    def safe_set_cell_with_fill(self, row, col, value, font=None, alignment=None, fill=None):
        """Safely set cell value with background fill (handles merged cells)"""
        try:
            cell = self.worksheet.cell(row=row, column=col)
            # Check if cell is part of a merged range
            if hasattr(cell, 'coordinate'):
                for merged_range in self.worksheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        # Unmerge the range first
                        self.worksheet.unmerge_cells(str(merged_range))
                        break
            
            # Now safely set the value
            cell = self.worksheet.cell(row=row, column=col)
            cell.value = value
            
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill
                
            return True
        except Exception as e:
            print(f"     ⚠️  Could not set cell with fill ({row}, {col}): {e}")
            return False

    def find_date_columns(self, date_row: int = 9, start_col: int = 3) -> Dict[str, int]:
        """
        Find which columns correspond to which dates.
        Auto-detects template type (schedule vs contact).
        
        Returns:
            Dict mapping date strings to column numbers
        """
        date_columns = {}
        
        # First, detect template type
        template_type = self._detect_template_type()
        print(f"🔍 Detected template type: {template_type}")
        
        if template_type == "contact":
            print("📞 Contact template detected - no date columns needed")
            return {}
        elif template_type == "schedule":
            return self._find_schedule_date_columns(date_row, start_col)
        else:
            print("❓ Unknown template type - trying schedule format...")
            return self._find_schedule_date_columns(date_row, start_col)
    
    def _find_schedule_date_columns(self, date_row: int = 9, start_col: int = 3) -> Dict[str, int]:
        """
        Find date columns for schedule templates.
        Row 8: Month names, Row 9: Day numbers, Row 10: Day symbols
        """
        date_columns = {}
        
        print(f"🗓️  Scanning row {date_row} for day numbers starting from column {start_col}...")
        
        # Get year and month context from row 8
        current_year = 2025  # Default year
        month_mapping = {}
        
        # Scan row 8 for month names to map columns to months
        for col in range(start_col, self.worksheet.max_column + 1):
            month_cell = self.worksheet.cell(row=8, column=col)
            if month_cell.value and isinstance(month_cell.value, str):
                month_name = month_cell.value.upper().strip()
                if 'SEPTEMBRE' in month_name:
                    month_mapping[col] = 9
                elif 'OCTOBRE' in month_name:
                    month_mapping[col] = 10
                elif 'NOVEMBRE' in month_name:
                    month_mapping[col] = 11
                elif 'DÉCEMBRE' in month_name:
                    month_mapping[col] = 12
                elif 'JANVIER' in month_name:
                    month_mapping[col] = 1
                elif 'FÉVRIER' in month_name:
                    month_mapping[col] = 2
                elif 'MARS' in month_name:
                    month_mapping[col] = 3
                elif 'AVRIL' in month_name:
                    month_mapping[col] = 4
                elif 'MAI' in month_name:
                    month_mapping[col] = 5
                elif 'JUIN' in month_name:
                    month_mapping[col] = 6
                elif 'JUILLET' in month_name:
                    month_mapping[col] = 7
                elif 'AOÛT' in month_name:
                    month_mapping[col] = 8
        
        # Now scan row 9 for day numbers
        current_month = 9  # Default to September
        for col in range(start_col, self.worksheet.max_column + 1):
            cell_value = self.worksheet.cell(row=date_row, column=col).value
            
            # Find the month for this column (look for nearest month header)
            for month_col, month_num in month_mapping.items():
                if col >= month_col:
                    current_month = month_num
            
            if cell_value:
                if isinstance(cell_value, (int, float)):
                    day = int(cell_value)
                    if 1 <= day <= 31:
                        date_str = f"{current_year}-{current_month:02d}-{day:02d}"
                        date_columns[date_str] = col
                        print(f"   📅 {date_str} (day {day}) → Column {col} ({chr(64 + col)})")
                elif isinstance(cell_value, datetime):
                    day = cell_value.day
                    date_str = f"{current_year}-{current_month:02d}-{day:02d}"
                    date_columns[date_str] = col
                    print(f"   📅 {date_str} (day {day}) → Column {col} ({chr(64 + col)})")
        
        print(f"✅ Found {len(date_columns)} date columns")
        return date_columns
    
    def find_team_member_rows(self, name_col: int = 2, start_row: int = 11) -> Dict[str, int]:
        """
        Find team member names and map them to their positions.
        For schedule templates, returns empty dict since we insert members dynamically.
        
        Returns:
            Dict mapping team member names to their row/position numbers
        """
        template_type = self._detect_template_type()
        
        if template_type == "contact":
            return self._find_contact_team_members()
        else:
            return self._find_schedule_team_members(name_col, start_row)
    
    def _find_contact_team_members(self) -> Dict[str, int]:
        """
        Find team members in contact template format.
        In template.xlsx: Row 15 has first names, Row 16 has last names
        """
        team_members = {}
        print("👥 Scanning for team members in contact template format...")
        
        # Check row 15 for first names
        first_names = {}
        for col in range(3, self.worksheet.max_column + 1, 3):  # Every 3rd column starting from C
            cell = self.worksheet.cell(row=15, column=col)
            if cell.value and isinstance(cell.value, str):
                first_name = cell.value.strip()
                if first_name:
                    first_names[col] = first_name
                    print(f"   👤 '{first_name}' → Column {col}")
        
        # For contact template, we'll use the first names as keys
        # and map them to a "virtual" row number for consistency
        row_num = 15  # Base row for contact template
        for col, name in first_names.items():
            team_members[name.title()] = row_num  # Use title case
        
        print(f"✅ Found {len(team_members)} team members in contact format")
        return team_members
    
    def _find_schedule_team_members(self, name_col: int = 2, start_row: int = 11) -> Dict[str, int]:
        """
        Find existing team members in the schedule template grid.
        Scans column B for team member names starting from the schedule area.
        """
        team_members = {}
        print(f"👥 Scanning column {name_col} for existing team members starting from row {start_row}...")
        
        # Scan the schedule grid area for team member names
        for row in range(start_row, min(25, self.worksheet.max_row + 1)):
            name_cell = self.worksheet.cell(row=row, column=name_col)
            if name_cell.value and isinstance(name_cell.value, str):
                name = name_cell.value.strip()
                # Skip headers and empty cells
                if name and name not in ['Sages femmes', 'SAGES FEMMES', '']:
                    team_members[name] = row
                    print(f"   👤 '{name}' → Row {row}")
        
        print(f"✅ Found {len(team_members)} existing team members in schedule grid")
        return team_members
    
    def fill_schedule_data_by_dates(self, schedule_data: Dict[str, Dict[str, str]]):
        """
        Fill in the schedule symbols using date-based mapping.
        Auto-detects template type and fills accordingly.
        
        Args:
            schedule_data: Dict mapping team member names to their schedules
                          Format: {'Member Name': {'2025-07-15': 'X', '2025-07-16': 'RP', ...}, ...}
        """
        print("📅 Filling schedule data using date-based mapping...")
        
        template_type = self._detect_template_type()
        
        if template_type == "contact":
            self._fill_contact_template_with_schedule(schedule_data)
        else:
            self._fill_schedule_template(schedule_data)
    
    def _fill_schedule_template(self, schedule_data: Dict[str, Dict[str, str]]):
        """
        Fill schedule data in existing schedule template grid.
        Maps team members to existing rows and fills their schedules.
        """
        print("📋 Filling existing schedule template grid...")
        
        # Get date-to-column mapping
        date_columns = self.find_date_columns()
        
        if not date_columns:
            print("❌ No date columns found - cannot fill schedule")
            return
        
        # Find existing team member rows
        team_member_rows = self.find_team_member_rows()
        
        if not team_member_rows:
            print("❌ No team member rows found in template")
            return
        
        filled_count = 0
        matched_members = 0
        
        # Fill schedule for each team member
        for member_name, member_schedule in schedule_data.items():
            # Try to find matching team member in template
            member_row = None
            
            # Look for exact match first
            if member_name in team_member_rows:
                member_row = team_member_rows[member_name]
            else:
                # Try partial matches (first name only)
                first_name = member_name.split()[0].upper()
                for template_name, row in team_member_rows.items():
                    if first_name in template_name.upper():
                        member_row = row
                        break
            
            if member_row:
                matched_members += 1
                print(f"   👤 Filling '{member_name}' at row {member_row}")
                
                # Fill schedule symbols for this member
                for date_str, symbol in member_schedule.items():
                    if date_str in date_columns:
                        col = date_columns[date_str]
                        
                        # Set the schedule symbol
                        cell = self.worksheet.cell(row=member_row, column=col)
                        cell.value = symbol
                        
                        # Apply Times New Roman font formatting
                        cell.font = Font(name='Times New Roman', size=11)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        filled_count += 1
                        print(f"     ✅ {date_str} = '{symbol}' (Row {member_row}, Col {col})")
                    else:
                        print(f"     ⚠️  Date {date_str} not found in template")
            else:
                print(f"   ❌ No matching row found for '{member_name}'")
        
        print(f"✅ Filled {matched_members}/{len(schedule_data)} team members with {filled_count} schedule entries")
    
    def _fill_contact_template_with_schedule(self, schedule_data: Dict[str, Dict[str, str]]):
        """
        For contact templates, create a schedule section below the contact info.
        """
        print("📋 Creating schedule section in contact template...")
        
        # Get member names from contact template
        member_rows = self.find_team_member_rows()
        
        if not member_rows:
            print("❌ No team members found in template")
            return
        
        # Create schedule headers starting after contact info (around row 20)
        schedule_start_row = 20
        
        # Add schedule title
        title_cell = self.worksheet.cell(row=schedule_start_row, column=2)
        title_cell.value = "HORAIRE / SCHEDULE"
        title_cell.font = Font(name='Times New Roman', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Collect all unique dates from schedule data
        all_dates = set()
        for member_schedule in schedule_data.values():
            all_dates.update(member_schedule.keys())
        
        sorted_dates = sorted(all_dates)
        print(f"   📅 Found {len(sorted_dates)} unique dates: {sorted_dates[:5]}...")
        
        # Create date headers starting from column C
        header_row = schedule_start_row + 2
        for i, date_str in enumerate(sorted_dates):
            col = 3 + i  # Start from column C
            cell = self.worksheet.cell(row=header_row, column=col)
            # Show just the day number
            day = date_str.split('-')[2]
            cell.value = int(day)
            cell.font = Font(name='Times New Roman', size=10, bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Add team members and their schedules
        data_start_row = header_row + 1
        filled_count = 0
        
        for i, (member_name, member_schedule) in enumerate(schedule_data.items()):
            member_row = data_start_row + i
            
            # Add member name in column B
            name_cell = self.worksheet.cell(row=member_row, column=2)
            name_cell.value = member_name
            name_cell.font = Font(name='Times New Roman', size=10)
            
            # Add schedule symbols
            for j, date_str in enumerate(sorted_dates):
                col = 3 + j
                symbol = member_schedule.get(date_str, '')
                if symbol:
                    cell = self.worksheet.cell(row=member_row, column=col)
                    cell.value = symbol
                    cell.font = Font(name='Times New Roman', size=10)
                    cell.alignment = Alignment(horizontal='center')
                    filled_count += 1
        
        print(f"✅ Added schedule section with {filled_count} entries")
    
    def save_filled_template(self, output_path: str):
        """
        Save the filled template to a new file.
        
        Args:
            output_path: Path where to save the filled template
        """
        try:
            self.workbook.save(output_path)
            print(f"💾 Filled template saved to: {output_path}")
        except Exception as e:
            print(f"❌ Error saving template: {e}")
            raise
    
    def fill_template_with_schedule_data(self, schedule_data: Dict[str, Dict[str, str]],
                                        contact_data: Optional[Dict[str, Dict[str, str]]] = None,
                                        output_path: str = "schedule_filled.xlsx"):
        """
        Complete workflow to fill the template with schedule data.
        
        Args:
            schedule_data: Schedule data in format:
                          {'Member Name': {'2025-07-15': 'X', '2025-07-16': 'RP', ...}, ...}
            contact_data: Optional contact information
            output_path: Where to save the filled template
        """
        print("🚀 Starting template filling process...")
        
        # Load template
        self.load_template()
        
        # Fill schedule data using date-based mapping
        self.fill_schedule_data_by_dates(schedule_data)
        
        # Save the result
        self.save_filled_template(output_path)
        
        print("✅ Template filling completed successfully!")
        return output_path

    @staticmethod
    def process_schedule_data_from_db(db_data: List[Dict]) -> Dict[str, Dict[str, str]]:
        """
        Convert database schedule data to the format expected by template filler.
        
        Args:
            db_data: List of records from database with fields like:
                    [{'name': 'John', 'date': '2025-07-15', 'symbol': 'X'}, ...]
        
        Returns:
            Schedule data in format:
            {'Member Name': {'2025-07-15': 'X', '2025-07-16': 'RP', ...}, ...}
        """
        schedule_dict = {}
        
        for record in db_data:
            member_name = record.get('name', '')
            date_str = record.get('date', '')
            symbol = record.get('symbol', '')
            
            if member_name not in schedule_dict:
                schedule_dict[member_name] = {}
            
            schedule_dict[member_name][date_str] = symbol
        
        return schedule_dict
