# backend/app.py

import sqlite3
from flask import Flask, jsonify, request
from flask_cors import CORS
from datetime import datetime, timedelta
import pandas as pd
import sys
import os
import traceback

try:
    import xlsxwriter
except ImportError:
    xlsxwriter = None
    print("Warning: xlsxwriter not available")

# Add the 'core' directory to the Python path to find our custom modules
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), 'core')))

# Import core modules
try:
    import processor
    import utils
    import exporter
except ImportError as e:
    print(f"Warning: Could not import core modules: {e}")
    processor = None
    utils = None
    exporter = None 

# --- App Setup ---
app = Flask(__name__)
CORS(app) 

# --- Database Configuration ---
def get_database_path():
    """Get database path that works in both development and packaged mode"""
    if getattr(sys, '_MEIPASS', None):
        # Running from PyInstaller bundle
        return os.path.join(sys._MEIPASS, 'core', 'data', 'workers.db')
    else:
        # Running from source
        return os.path.join(os.path.dirname(__file__), 'core', 'data', 'workers.db')

DATABASE_PATH = get_database_path()

def get_archive_directory():
    """Get the standardized archive directory path"""
    backend_dir = os.path.dirname(__file__)
    archive_dir = os.path.join(backend_dir, 'core', 'data', 'archive')
    os.makedirs(archive_dir, exist_ok=True)
    return archive_dir

def get_db_connection():
    # Ensure the directory exists
    db_dir = os.path.dirname(DATABASE_PATH)
    os.makedirs(db_dir, exist_ok=True)
    
    conn = sqlite3.connect(DATABASE_PATH, timeout=30.0)
    conn.row_factory = sqlite3.Row 
    # Enable WAL mode for better concurrency
    conn.execute('PRAGMA journal_mode=WAL')
    return conn

# Helper function to ensure database schema is up to date
def migrate_database():
    """Ensure database schema has all required columns"""
    conn = None
    try:
        conn = get_db_connection()
        
        # Create the table with complete schema
        conn.execute('''
            CREATE TABLE IF NOT EXISTS employee_schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                date DATE,
                day_of_week INTEGER,
                start_time TEXT,
                end_time TEXT,
                schedule_type TEXT DEFAULT 'work',
                activity_type TEXT DEFAULT 'X',
                is_night_shift BOOLEAN DEFAULT 0,
                shift_number INTEGER DEFAULT 1,
                hours_worked REAL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES workers (id)
            )
        ''')
        
        # Add missing columns if they don't exist
        columns_to_add = [
            ('activity_type', 'TEXT DEFAULT "X"'),
            ('shift_number', 'INTEGER DEFAULT 1'),
            ('is_night_shift', 'BOOLEAN DEFAULT 0'),
            ('hours_worked', 'REAL DEFAULT 0'),
            ('updated_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
        ]
        
        for column_name, column_def in columns_to_add:
            try:
                conn.execute(f'ALTER TABLE employee_schedules ADD COLUMN {column_name} {column_def}')
                print(f"Added column {column_name}")
            except sqlite3.OperationalError as e:
                if "duplicate column name" in str(e):
                    pass  # Column already exists
                else:
                    print(f"Error adding column {column_name}: {e}")
        
        conn.commit()
        
        # Create teams management tables
        create_teams_tables(conn)
        
        # Ensure workers table exists with phone_number column
        create_workers_table(conn)
        
    except Exception as e:
        print(f"Error migrating database: {e}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()

def create_workers_table(conn):
    """Create workers table with all required columns including phone_number"""
    try:
        # Create workers table if it doesn't exist
        conn.execute('''
            CREATE TABLE IF NOT EXISTS workers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                status TEXT DEFAULT 'employee',
                employee_number TEXT,
                phone_number TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Add missing columns if they don't exist
        columns_to_add = [
            ('employee_number', 'TEXT'),
            ('phone_number', 'TEXT'),
            ('created_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP'),
            ('updated_at', 'TIMESTAMP DEFAULT CURRENT_TIMESTAMP')
        ]
        
        for column_name, column_def in columns_to_add:
            try:
                conn.execute(f'ALTER TABLE workers ADD COLUMN {column_name} {column_def}')
                print(f"Added column {column_name} to workers table")
            except sqlite3.OperationalError as e:
                if "duplicate column name" in str(e):
                    pass  # Column already exists
                else:
                    print(f"Error adding column {column_name}: {e}")
        
        conn.commit()
        print("Workers table created/updated successfully")
        
    except Exception as e:
        print(f"Error creating workers table: {e}")
        conn.rollback()

def create_teams_tables(conn):
    """Create tables for teams management"""
    try:
        # Teams table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS teams (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                description TEXT,
                color TEXT DEFAULT '#3b82f6',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Team members table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS team_members (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                team_id INTEGER,
                employee_id INTEGER,
                role TEXT DEFAULT 'member',
                joined_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (team_id) REFERENCES teams (id),
                FOREIGN KEY (employee_id) REFERENCES workers (id),
                UNIQUE(team_id, employee_id)
            )
        ''')
        
        # Team shifts table
        conn.execute('''
            CREATE TABLE IF NOT EXISTS team_shifts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                team_id INTEGER,
                name TEXT NOT NULL,
                start_time TEXT NOT NULL,
                end_time TEXT NOT NULL,
                frequency INTEGER DEFAULT 1,
                start_date DATE NOT NULL,
                end_date DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (team_id) REFERENCES teams (id)
            )
        ''')
        
        conn.commit()
        print("Teams tables created successfully")
        
    except Exception as e:
        print(f"Error creating teams tables: {e}")
        conn.rollback()

# --- API Endpoints ---
@app.route('/api/workers', methods=['GET'])
def get_workers():
    print("Request received for /api/workers")
    
    # Get query parameters for filtering
    search_query = request.args.get('search', '').strip()
    status_filter = request.args.get('status', '').strip()
    
    print(f"Filters - search: '{search_query}', status: '{status_filter}'")
    
    conn = get_db_connection()
    
    # Check if employee_number column exists
    cursor = conn.cursor()
    cursor.execute('PRAGMA table_info(workers)')
    columns = [row[1] for row in cursor.fetchall()]
    
    # Build the SQL query with filters
    base_query = "SELECT id, name, status"
    if 'employee_number' in columns:
        base_query += ", employee_number"
    else:
        base_query += ", NULL as employee_number"
    if 'phone_number' in columns:
        base_query += ", phone_number"
    else:
        base_query += ", NULL as phone_number"
    
    base_query += " FROM workers"
    
    # Add WHERE conditions for filtering
    where_conditions = []
    params = []
    
    if search_query:
        # Search in name, id, and employee_number (if exists)
        search_conditions = ["name LIKE ?", "CAST(id AS TEXT) LIKE ?"]
        if 'employee_number' in columns:
            search_conditions.append("employee_number LIKE ?")
        
        where_conditions.append(f"({' OR '.join(search_conditions)})")
        search_param = f"%{search_query}%"
        params.extend([search_param] * len(search_conditions))
    
    if status_filter:
        where_conditions.append("status = ?")
        params.append(status_filter)
    
    # Combine query with filters
    if where_conditions:
        base_query += " WHERE " + " AND ".join(where_conditions)
    
    base_query += " ORDER BY name ASC"
    
    print(f"Executing query: {base_query}")
    print(f"Parameters: {params}")
    
    workers_cursor = conn.execute(base_query, params).fetchall()
    conn.close()
    
    workers_list = [dict(row) for row in workers_cursor]
    print(f"Found {len(workers_list)} workers matching filters")
    
    return jsonify(workers_list)

# NEW ENDPOINT for weekly schedule
@app.route('/api/schedule/weekly', methods=['GET'])
def get_weekly_schedule():
    # Get the date from the query parameter, default to today if not provided
    date_str = request.args.get('date', default=datetime.today().strftime('%Y-%m-%d'))
    print(f"Request received for /api/schedule/weekly for date: {date_str}")
    
    try:
        week_start_dt = datetime.strptime(date_str, '%Y-%m-%d')
        
        # Find the Sunday of that week (start of week)
        day_of_week = week_start_dt.weekday()  # Monday = 0, Sunday = 6
        days_to_subtract = (day_of_week + 1) % 7  # Convert to Sunday = 0 format
        week_start = week_start_dt - timedelta(days=days_to_subtract)
        week_end = week_start + timedelta(days=6)
        
        print(f"Week range: {week_start.date()} to {week_end.date()}")
        
        # Get all workers
        conn = get_db_connection()
        workers = conn.execute('SELECT id, name FROM workers ORDER BY name ASC').fetchall()
        
        # Get all schedules for this week (including night shift end markers)
        schedules = conn.execute('''
            SELECT employee_id, date, start_time, end_time, activity_type, is_night_shift, hours_worked, schedule_type
            FROM employee_schedules
            WHERE date >= ? AND date <= ?
            ORDER BY employee_id, date
        ''', (week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d'))).fetchall()
        
        conn.close()
        
        # Build schedule data for each worker
        schedule_data = []
        
        for worker in workers:
            worker_id = worker['id']
            worker_name = worker['name']
            
            # Initialize empty week
            week_schedule = {
                'Name': worker_name,
                'Sunday': '',
                'Monday': '',
                'Tuesday': '',
                'Wednesday': '',
                'Thursday': '',
                'Friday': '',
                'Saturday': '',
                'TOTAL': 0
            }
            
            # Fill in schedules for this worker
            worker_schedules = [s for s in schedules if s['employee_id'] == worker_id]
            total_hours = 0
            work_days = 0
            
            # Track which days have work schedules (not night shift end markers)
            working_days = set()
            
            for schedule in worker_schedules:
                schedule_date = datetime.strptime(schedule['date'], '%Y-%m-%d').date()
                day_of_week = schedule_date.weekday()  # Monday = 0, Sunday = 6
                
                # Convert to our day names
                day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                day_name = day_names[day_of_week]
                
                # Handle different schedule types
                if schedule['schedule_type'] == 'work':
                    # Normal work schedule
                    activity = schedule['activity_type'] or 'x'
                    
                    # Fix the activity symbols to match your requirements
                    if activity.upper() == 'X':
                        display_value = 'x'  # Working (lowercase)
                    elif activity.upper() == 'S':
                        display_value = 'S'  # External person helping
                    elif activity.upper() == 'RP':
                        display_value = 'RP'  # Meeting
                    elif activity.upper() == 'M':
                        display_value = 'M'  # Maladie (sick leave)
                    elif activity.upper() == 'V':
                        display_value = 'V'  # Vacances (vacation)
                    else:
                        display_value = activity.lower()  # Default to lowercase
                    
                    # Add to total hours and track working day (exclude V and M from totals)
                    if activity.upper() not in ['V', 'M']:  # Don't count Vacances (V) or Maladie (M)
                        total_hours += schedule['hours_worked'] or 0
                        working_days.add(schedule_date)
                    
                    # For night shifts, also count the next day (but only if it's within this week)
                    if schedule['is_night_shift'] and activity.upper() not in ['V', 'M']:
                        next_day = schedule_date + timedelta(days=1)
                        # Only count the next day if it falls within the current week
                        if week_start.date() <= next_day <= week_end.date():
                            working_days.add(next_day)
                    
                elif schedule['schedule_type'] == 'unavailable':
                    # Unavailability (M or V) - do NOT count in totals
                    activity = schedule['activity_type'] or 'M'
                    display_value = activity.upper()  # M or V
                    # Note: we deliberately do NOT add to working_days or total_hours
                    
                elif schedule['schedule_type'] == 'end_of_night_shift':
                    # Night shift ending - show F + time only if end time is before 12:00 PM (noon)
                    # Otherwise, show as regular X shift
                    # Only count working days for non-V and non-M activities
                    activity = schedule['activity_type'] if schedule['activity_type'] else 'X'
                    if activity.upper() not in ['V', 'M']:
                        working_days.add(schedule_date)  # Count the end day of night shift
                    
                    end_time = schedule['end_time']
                    if end_time:
                        # Check if end time is before 12:00 PM (noon)
                        hour = int(end_time.split(':')[0])
                        if hour < 12:  # Before noon, use F\n format
                            display_value = f"F\n{hour}h"
                        else:  # At or after noon, use X format
                            display_value = "X"
                    else:
                        display_value = "F"
                else:
                    continue  # Skip unknown schedule types
                
                # If there's already something for this day, check if it's a duplicate
                if week_schedule[day_name]:
                    # Check if this value is already present to avoid duplicates
                    existing_values = week_schedule[day_name].split(", ")
                    if display_value not in existing_values:
                        week_schedule[day_name] += f", {display_value}"
                else:
                    week_schedule[day_name] = display_value
            
            # Count working days and format total as "X (YH)"
            work_days = len(working_days)
            if work_days > 0 and total_hours > 0:
                week_schedule['TOTAL'] = f"{work_days} ({total_hours:.0f}H)"
            elif work_days > 0:
                week_schedule['TOTAL'] = f"{work_days} (0H)"
            else:
                week_schedule['TOTAL'] = "0"
            
            schedule_data.append(week_schedule)
        
        print(f"Generated schedule for {len(schedule_data)} workers")
        return jsonify(schedule_data)
            
    except Exception as e:
        print(f"Error generating weekly schedule: {e}")
        import traceback
        traceback.print_exc()
        
        # Return empty data instead of sample data for new users
        return jsonify([])

# NEW ENDPOINT for generating weekly reports
@app.route('/api/reports/weekly', methods=['POST'])
def generate_weekly_report():
    try:
        data = request.get_json()
        date_str = data.get('date', datetime.today().strftime('%Y-%m-%d'))
        week_start = datetime.strptime(date_str, '%Y-%m-%d')
        
        # Generate the report using existing processor functionality
        processor.currentWeekRepport()
        
        return jsonify({"message": "Weekly report generated successfully", "date": date_str})
    except Exception as e:
        print(f"Error generating weekly report: {e}")
        return jsonify({"error": str(e)}), 500

# NEW ENDPOINT for Excel export
@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    try:
        date_str = request.args.get('date', default=datetime.today().strftime('%Y-%m-%d'))
        format_type = request.args.get('format', 'standard')  # 'standard' or 'professional'
        selected_date = datetime.strptime(date_str, '%Y-%m-%d')
        
        # Calculate the actual week start (Sunday)
        day_of_week = selected_date.weekday()  # Monday = 0, Sunday = 6
        days_to_subtract = (day_of_week + 1) % 7  # Convert to Sunday = 0 format
        week_start = selected_date - timedelta(days=days_to_subtract)
        week_start_str = week_start.strftime('%Y-%m-%d')
        
        print(f"Export request for date: {date_str}, week_start: {week_start_str}, format: {format_type}")
        
        # If professional format is requested, use the new exporter
        if format_type == 'professional' and exporter:
            # Calculate end date for the week
            week_end = week_start + timedelta(days=6)
            
            # Get schedule data similar to the professional export endpoint
            conn = get_db_connection()
            workers = conn.execute('SELECT id, name FROM workers ORDER BY name ASC').fetchall()
            schedules = conn.execute('''
                SELECT employee_id, date, start_time, end_time, activity_type, is_night_shift, 
                       hours_worked, schedule_type
                FROM employee_schedules
                WHERE date >= ? AND date <= ?
                ORDER BY employee_id, date
            ''', (week_start_str, week_end.strftime('%Y-%m-%d'))).fetchall()
            conn.close()
            
            # Build schedule data
            schedule_data = {}
            for worker in workers:
                worker_id = worker['id']
                worker_name = worker['name']
                worker_schedules = [s for s in schedules if s['employee_id'] == worker_id]
                
                worker_data = {'employee_name': worker_name, 'employee_id': worker_id, 'schedules': {}}
                
                for schedule in worker_schedules:
                    schedule_date = schedule['date']
                    if schedule['schedule_type'] == 'work':
                        activity = schedule['activity_type'] or 'X'
                        display_value = activity.upper()
                    elif schedule['schedule_type'] == 'end_of_night_shift':
                        display_value = 'F'
                    else:
                        continue
                    
                    worker_data['schedules'][schedule_date] = {
                        'activity_type': display_value,
                        'start_time': schedule['start_time'],
                        'end_time': schedule['end_time'],
                        'hours_worked': schedule['hours_worked'],
                        'is_night_shift': schedule['is_night_shift']
                    }
                
                schedule_data[worker_name] = worker_data
            
            # Use professional exporter
            file_path = exporter.generate_professional_schedule_export(
                start_date=week_start,
                end_date=week_end,
                schedule_data=schedule_data
            )
            filename = os.path.basename(file_path)
            
            return jsonify({
                "message": f"Rapport Excel professionnel sauvegardé avec succès dans l'archive",
                "filename": filename,
                "path": file_path,
                "format": "professional"
            })
        
        # Standard format (existing logic)
        # Create archive directory path
        archive_dir = get_archive_directory()
        
        # Create filename using week start date
        filename = f'Rapport_Hebdomadaire_{week_start_str}.xlsx'
        file_path = os.path.join(archive_dir, filename)
        
        print(f"Saving Excel file to: {file_path}")
        
        # Get real data from database
        conn = get_db_connection()
        
        # Check what columns exist in the workers table
        cursor = conn.cursor()
        cursor.execute('PRAGMA table_info(workers)')
        columns = [row[1] for row in cursor.fetchall()]
        print(f"Available columns in workers table: {columns}")
        
        # Adjust query based on available columns
        if 'first_name' in columns and 'last_name' in columns:
            employees_query = """
            SELECT id, first_name, last_name, employee_number 
            FROM workers 
            ORDER BY last_name, first_name
            """
        else:
            # Fallback to name column
            employees_query = """
            SELECT id, name, employee_number 
            FROM workers 
            ORDER BY name
            """
        
        employees_df = pd.read_sql_query(employees_query, conn)
        print(f"Retrieved {len(employees_df)} employees")
        
        # Calculate the week dates
        week_dates = []
        for i in range(7):  # Sunday to Saturday
            day_date = week_start + timedelta(days=i)
            week_dates.append(day_date)
        
        # Get schedules for this week
        schedules_query = """
        SELECT employee_id, date, day_of_week, activity_type, start_time, end_time, is_night_shift
        FROM employee_schedules 
        WHERE date BETWEEN ? AND ?
        """
        end_date = week_start + timedelta(days=6)
        
        try:
            schedules_df = pd.read_sql_query(schedules_query, conn, params=[week_start.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d')])
            print(f"Retrieved {len(schedules_df)} schedule entries")
        except Exception as e:
            print(f"Error querying schedules: {e}")
            schedules_df = pd.DataFrame()  # Empty DataFrame as fallback
        
        conn.close()
        
        # Create the weekly schedule data structure
        schedule_data = []
        
        for _, employee in employees_df.iterrows():
            employee_id = employee['id']
            
            # Handle name field based on available columns
            if 'first_name' in employee and 'last_name' in employee:
                name = f"{employee['first_name']} {employee['last_name']}"
            else:
                name = employee['name']
            
            # Get employee's schedules for the week
            emp_schedules = schedules_df[schedules_df['employee_id'] == employee_id] if not schedules_df.empty else pd.DataFrame()
            
            # Initialize row data
            row_data = {"Name": name}
            day_names = ['Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi']
            total_hours = 0
            
            # Fill in schedule for each day
            for day_idx, day_name in enumerate(day_names):
                if not emp_schedules.empty:
                    day_schedule = emp_schedules[emp_schedules['day_of_week'] == day_idx]
                    
                    if not day_schedule.empty:
                        # Get the first schedule for this day (assuming one schedule per day)
                        schedule = day_schedule.iloc[0]
                        activity = schedule['activity_type'] if pd.notna(schedule['activity_type']) and schedule['activity_type'] else ''
                        
                        # Add time if it's a work day
                        if activity and activity != '':
                            start_time = schedule['start_time'] if pd.notna(schedule['start_time']) and schedule['start_time'] else ''
                            end_time = schedule['end_time'] if pd.notna(schedule['end_time']) and schedule['end_time'] else ''
                            
                            # Format the cell content
                            if pd.notna(schedule['is_night_shift']) and schedule['is_night_shift']:
                                row_data[day_name] = f"{activity}"  # Night shifts might have different formatting
                            elif start_time and end_time and activity == 'F':
                                # Apply proper F formatting logic
                                end_hour = int(end_time.split(':')[0])
                                if end_hour < 12:  # Before noon, use F format
                                    row_data[day_name] = f"F{end_time}"  # Format like "F09:00"
                                else:  # At or after noon, use X format
                                    row_data[day_name] = "X"
                            else:
                                row_data[day_name] = activity
                            
                            # Calculate hours for total (exclude M and V from totals)
                            if activity in ['X', 'S', 'RP', 'F', 'C'] and activity not in ['M', 'V']:
                                total_hours += 1  # Count work days, not hours
                        else:
                            row_data[day_name] = ''
                    else:
                        row_data[day_name] = ''
                else:
                    row_data[day_name] = ''
            
            row_data['TOTAL'] = total_hours
            schedule_data.append(row_data)
        
        # If no real data, return empty dataframe
        if not schedule_data:
            print("No schedule data found, returning empty dataframe")
            # Create empty dataframe with proper columns
            schedule_data = []
        
        df = pd.DataFrame(schedule_data)
        print(f"Created DataFrame with {len(df)} employees")
        
        # Save Excel file directly to archive directory
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Horaires_Equipes', index=False, header=False, startrow=9, startcol=1)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Horaires_Equipes']
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 20  # Name column
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    worksheet.column_dimensions[col].width = 12
        except ImportError as e:
            print(f"openpyxl not available, using xlsxwriter: {e}")
            # Fallback to xlsxwriter if openpyxl is not available
            df.to_excel(file_path, sheet_name='Horaires_Equipes', index=False, engine='xlsxwriter')
        
        # Verify file was created
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
            print(f"Excel file saved to archive: {file_path}")
            print(f"   File size: {file_size} bytes")
        else:
            raise Exception(f"File was not created: {file_path}")
        
        # Return success message only - no file download
        return jsonify({
            "message": f"Rapport Excel sauvegardé avec succès dans l'archive",
            "filename": filename,
            "path": file_path,
            "employees_count": len(df)
        })
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        traceback.print_exc()
        
        # More detailed error information
        error_details = {
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()
        }
        print(f"Detailed error: {error_details}")
        
        return jsonify({"error": f"Export failed: {str(e)}"}), 500

# NEW ENDPOINT for Professional Schedule Export (modelExcel.xlsx format)
@app.route('/api/export/professional-schedule', methods=['POST'])
def export_professional_schedule():
    """Export schedules in professional format matching modelExcel.xlsx"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "Request body is required"}), 400
        
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        
        if not start_date_str or not end_date_str:
            return jsonify({"error": "start_date and end_date parameters are required"}), 400
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
        
        print(f"Professional export request: {start_date_str} to {end_date_str}")
        
        # Get schedule data from database
        conn = get_db_connection()
        
        # Get all workers
        workers = conn.execute('SELECT id, name FROM workers ORDER BY name ASC').fetchall()
        
        # Get all schedules for the date range
        schedules = conn.execute('''
            SELECT employee_id, date, start_time, end_time, activity_type, is_night_shift, 
                   hours_worked, schedule_type
            FROM employee_schedules
            WHERE date >= ? AND date <= ?
            ORDER BY employee_id, date
        ''', (start_date_str, end_date_str)).fetchall()
        
        conn.close()
        
        # Build schedule data for the professional exporter
        schedule_data = {}
        
        for worker in workers:
            worker_id = worker['id']
            worker_name = worker['name']
            
            # Get schedules for this worker
            worker_schedules = [s for s in schedules if s['employee_id'] == worker_id]
            
            # Build date-keyed schedule data
            worker_data = {
                'employee_name': worker_name,
                'employee_id': worker_id,
                'schedules': {}
            }
            
            for schedule in worker_schedules:
                schedule_date = schedule['date']
                
                # Handle different schedule types
                if schedule['schedule_type'] == 'work':
                    # Normal work schedule
                    activity = schedule['activity_type'] or 'X'
                    
                    # Normalize activity types
                    if activity.upper() == 'X':
                        display_value = 'X'
                    elif activity.upper() == 'S':
                        display_value = 'S'  
                    elif activity.upper() == 'RP':
                        display_value = 'RP'
                    elif activity.upper() == 'V':
                        display_value = 'V'  # Vacances
                    elif activity.upper() == 'C':
                        display_value = 'C'  # Congé
                    elif activity.upper() == 'M':
                        display_value = 'M'  # Maladie
                    else:
                        display_value = activity.upper()
                
                elif schedule['schedule_type'] == 'end_of_night_shift':
                    # Night shift ending - show F
                    display_value = 'F'
                else:
                    continue  # Skip unknown schedule types
                
                # Store the schedule data
                worker_data['schedules'][schedule_date] = {
                    'activity_type': display_value,
                    'start_time': schedule['start_time'],
                    'end_time': schedule['end_time'],
                    'hours_worked': schedule['hours_worked'],
                    'is_night_shift': schedule['is_night_shift']
                }
            
            schedule_data[worker_name] = worker_data
        
        print(f"Built schedule data for {len(schedule_data)} workers")
        
        # Use the professional exporter
        if exporter:
            file_path = exporter.generate_professional_schedule_export(
                start_date=start_date,
                end_date=end_date,
                schedule_data=schedule_data
            )
            
            filename = os.path.basename(file_path)
            
            # Verify file was created
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                print(f"Professional schedule export saved: {file_path}")
                print(f"   File size: {file_size} bytes")
                
                return jsonify({
                    "message": f"Horaire professionnel exporté avec succès dans l'archive",
                    "filename": filename,
                    "path": file_path,
                    "date_range": f"{start_date_str} au {end_date_str}",
                    "format": "professional"
                })
            else:
                raise Exception(f"File was not created: {file_path}")
        else:
            return jsonify({"error": "Professional exporter module not available"}), 500
        
    except Exception as e:
        print(f"Error exporting professional schedule: {e}")
        traceback.print_exc()
        
        error_details = {
            "error": str(e),
            "error_type": type(e).__name__,
            "traceback": traceback.format_exc()
        }
        print(f"Detailed error: {error_details}")
        
        return jsonify({"error": f"Professional export failed: {str(e)}"}), 500


# NEW ENDPOINT for Team Schedules Excel export
@app.route('/api/export/team-schedules', methods=['POST'])
def export_team_schedules():
    """Export team schedules to Excel format using the exact same data as the UI"""
    try:
        # Get request data
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "Request body is required"}), 400
            
        start_date_str = data.get('start_date')
        end_date_str = data.get('end_date')
        teams_data = data.get('teams', [])
        
        if not start_date_str or not end_date_str:
            return jsonify({"error": "start_date and end_date parameters are required"}), 400
        
        if not teams_data:
            return jsonify({"error": "teams data is required"}), 400
        
        print(f"Export request: {start_date_str} to {end_date_str}")
        
        # Parse dates
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # Get schedule data using the same logic as team-range endpoint
        conn = get_db_connection()
        
        # Get all workers
        workers = conn.execute('SELECT id, name FROM workers ORDER BY name ASC').fetchall()
        
        # Get all schedules for the date range
        schedules = conn.execute('''
            SELECT employee_id, date, start_time, end_time, activity_type, is_night_shift, 
                   hours_worked, schedule_type
            FROM employee_schedules
            WHERE date >= ? AND date <= ?
            ORDER BY employee_id, date
        ''', (start_date_str, end_date_str)).fetchall()
        
        conn.close()
        
        # Build schedule data for each worker
        worker_schedule_data = {}
        
        # First, initialize all workers with empty schedule data
        for worker in workers:
            worker_id = worker['id']
            worker_name = worker['name']
            worker_schedule_data[worker_name] = {
                'employee_name': worker_name,
                'employee_id': worker_id,
                'schedules': {}
            }
        
        # Then, populate schedules for workers who have them
        for worker in workers:
            worker_id = worker['id']
            worker_name = worker['name']
            worker_schedules = [s for s in schedules if s['employee_id'] == worker_id]
            
            for schedule in worker_schedules:
                schedule_date = schedule['date']
                display_value = ''
                if schedule['schedule_type'] == 'work':
                    activity = schedule['activity_type'] or 'X'
                    start_time = schedule['start_time']
                    if activity.upper() == 'X':
                        if start_time and start_time >= '16:00':
                            hour = start_time.split(':')[0]
                            display_value = f"D\n{hour}h"
                        else:
                            display_value = 'X'
                    elif activity.upper() in ['S', 'RP', 'V', 'C', 'M']:
                        display_value = activity.upper()
                    elif activity.upper() == 'D':
                        if start_time:
                            hour = start_time.split(':')[0]
                            display_value = f"D\n{hour}h"
                        else:
                            display_value = "D\n17h"
                    else:
                        if start_time and start_time >= '16:00':
                            hour = start_time.split(':')[0]
                            display_value = f"D\n{hour}h"
                        else:
                            display_value = activity.upper()
                elif schedule['schedule_type'] == 'end_of_night_shift':
                    end_time = schedule['end_time']
                    if end_time:
                        hour = int(end_time.split(':')[0])
                        if hour < 12:
                            display_value = f"F\n{hour}h"
                        else:
                            display_value = "X"
                    else:
                        display_value = "F\n9h"
                else:
                    continue
                
                worker_schedule_data[worker_name]['schedules'][schedule_date] = {
                    'activity_type': display_value,
                    'start_time': schedule['start_time'],
                    'end_time': schedule['end_time'],
                    'hours_worked': schedule['hours_worked'],
                    'is_night_shift': schedule['is_night_shift']
                }
        
        # Generate date range
        date_range = [start_date + timedelta(days=i) for i in range((end_date - start_date).days + 1)]
        
        # Create column headers for the DataFrame
        column_headers = ['Sage-femme'] + [d.strftime('%Y-%m-%d') for d in date_range]

        # Create data rows
        data_rows = []
        for team_index, team in enumerate(teams_data):
            # Get team members - use the member data directly from the frontend
            team_members = team.get('members', [])
            
            if team_members:
                for member in team_members:
                    member_name = member.get('name', '')
                    if not member_name:
                        continue
                    
                    # Try to find schedule data for this member by name (same logic as frontend)
                    member_schedule_data = None
                    
                    # Strategy 1: Match by exact name
                    if member_name in worker_schedule_data:
                        member_schedule_data = worker_schedule_data[member_name]
                    else:
                        # Strategy 2: Match by partial name (first name or last name)
                        member_name_parts = member_name.lower().split(' ')
                        for worker_name, worker_data in worker_schedule_data.items():
                            worker_name_parts = worker_name.lower().split(' ')
                            # Check if any part of the names match
                            if any(part in worker_name_parts for part in member_name_parts if len(part) > 2):
                                member_schedule_data = worker_data
                                print(f"Matched {member_name} with {worker_name} by partial name")
                                break
                    
                    if not member_schedule_data:
                        print(f"No schedule data found for team member: {member_name}")
                        # Create empty schedule data for this member
                        member_schedule_data = {
                            'employee_name': member_name,
                            'schedules': {}
                        }
                    
                    # Build row data for this member
                    row_data = {'Sage-femme': member_name}
                    
                    for date in date_range:
                        date_str = date.strftime('%Y-%m-%d')
                        schedule_info = member_schedule_data['schedules'].get(date_str)
                        
                        if schedule_info:
                            activity = schedule_info['activity_type']
                            # Handle special symbols for Excel export
                            if activity == 'M':  # Maladie/Sick - empty cell
                                row_data[date_str] = ''
                            elif activity == 'V':  # Vacances/Vacation - dash
                                row_data[date_str] = '-'
                            else:
                                row_data[date_str] = activity
                        else:
                            # Apply F\n9h injection logic - same as frontend team schedules
                            # Check if previous day was a night shift
                            prev_date = date - timedelta(days=1)
                            prev_date_str = prev_date.strftime('%Y-%m-%d')
                            prev_schedule = member_schedule_data['schedules'].get(prev_date_str)
                            
                            if prev_schedule and prev_schedule.get('is_night_shift'):
                                # Inject F\n9h using the end time from previous day, default to 9h
                                end_time = prev_schedule.get('end_time', '09:00')
                                if end_time:
                                    try:
                                        hour = int(end_time.split(':')[0])
                                        if hour < 12:  # Morning end time, use F\nXh format
                                            row_data[date_str] = f"F\n{hour}h"
                                        else:
                                            row_data[date_str] = ''
                                    except (ValueError, IndexError):
                                        row_data[date_str] = "F\n9h"  # Default fallback
                                else:
                                    row_data[date_str] = "F\n9h"  # Default fallback
                            else:
                                row_data[date_str] = ''
                    
                    data_rows.append(row_data)

        df = pd.DataFrame(data_rows, columns=column_headers)
        
        # Generate filename
        base_filename = f'Horaires_Equipes_{start_date_str}_au_{end_date_str}'
        archive_dir = get_archive_directory()
        counter = 0
        while True:
            filename = f'{base_filename}{f"_{counter}" if counter else ""}.xlsx'
            file_path = os.path.join(archive_dir, filename)
            if not os.path.exists(file_path):
                break
            try:
                with open(file_path, 'r+b'):
                    break
            except (PermissionError, IOError):
                counter += 1
                if counter > 100:
                    raise Exception("Unable to find available filename after 100 attempts")

        # Create Excel file
        try:
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
            import csv # Moved import here to fix NameError

            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Horaires_Equipes', index=False, header=False, startrow=9, startcol=1)
                
                workbook = writer.book
                worksheet = writer.sheets['Horaires_Equipes']

                # Page setup for easy printing
                worksheet.page_setup.orientation = 'landscape'
                worksheet.page_setup.fitToPage = True
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 1
                
                worksheet.page_margins.left = 0.0
                worksheet.page_margins.right = 0.15
                worksheet.page_margins.top = 0.15
                worksheet.page_margins.bottom = 0.15
                worksheet.page_margins.header = 0.0
                worksheet.page_margins.footer = 0.0

                # --- HEADER FORMATTING (MODIFICATIONS START) ---
                # Define styles
                left_alignment_no_wrap = Alignment(horizontal='left', vertical='center', wrap_text=False)
                franklin_gothic_medium_font = Font(name='Franklin Gothic Medium')
                franklin_gothic_book_font_9 = Font(name='Franklin Gothic Book', size=9)
                franklin_gothic_book_font_11_bold = Font(name='Franklin Gothic Book', size=11, bold=True)
                franklin_gothic_book_font_12_bold = Font(name='Franklin Gothic Book', size=12, bold=True)
                schedule_font = Font(size=9, bold=True)
                thick_black_bottom_border = Border(bottom=Side(border_style="thick", color="000000"))

                # Apply default font and alignment for the first 5 rows
                for row in range(2, 6): # Start from row 2
                    for col in range(1, 50):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = franklin_gothic_book_font_9 # Default to size 9
                        cell.alignment = left_alignment_no_wrap

                # Row 1: "Horaire de garde des sages femmes" - Override font
                worksheet.merge_cells('B1:X1')
                header_cell = worksheet['B1']
                header_cell.value = 'Horaire de garde des sages femmes'
                header_cell.font = Font(name='Franklin Gothic Medium', size=28)

                # Year formula in AI1:AL1 - Override font
                worksheet.merge_cells('AI1:AL1')
                year_cell = worksheet['AI1']
                year_cell.value = '=IF(MONTH(TODAY())=12,YEAR(TODAY())+1,YEAR(TODAY()))'
                year_cell.font = Font(name='Franklin Gothic Medium', bold=True, size=26) # Keep consistent with row font

                # Apply thick bottom border from B1 to AL1
                for col_idx in range(2, 39):  # From column B to AL1
                    worksheet.cell(row=1, column=col_idx).border = thick_black_bottom_border


                # Row 2: Maison de naissance - Override font
                worksheet.merge_cells('B2:J2')
                contact_left = worksheet['B2']
                contact_left.value = 'Maison de naissance de la Rivière'
                contact_left.font = franklin_gothic_book_font_12_bold # Override to size 12 bold

                # Add new merged cells for Row 2
                worksheet.merge_cells('AA2:AC2')
                legend_x_garde = worksheet['AA2']
                legend_x_garde.value = 'X - Garde'

                worksheet.merge_cells('AD2:AJ2')
                legend_s_soutien = worksheet['AD2']
                legend_s_soutien.value = 'S - Soutien de jour seulement'

                # Row 3: Address and Legends
                worksheet.merge_cells('B3:J3') # Extended merge
                address_cell = worksheet['B3']
                address_cell.value = '1275, rue Saint-Jean-Baptiste, Nicolet (Québec) J3T 1W4'

                worksheet.merge_cells('L3:Q3') # Moved Fax
                fax_cell = worksheet['L3']
                fax_cell.value = 'Fax : 819 293-2804'
                
                worksheet.merge_cells('AA3:AF3') # Changed from AA3:AF4
                legend_rp = worksheet['AA3']
                legend_rp.value = 'RP - Rencontre Prénatale'

                worksheet.merge_cells('AG3:AI3') # Changed from AG3:AI3
                conge_cell = worksheet['AG3']
                conge_cell.value = ' - : Congés'

                # Row 4: Phone numbers and other legends
                worksheet.merge_cells('B4:E4')
                phone_cell = worksheet['B4']
                phone_cell.value = 'Tél. : 819 293-2071, poste 56221'
                
                worksheet.merge_cells('G4:Q4')
                phone_sans_frais = worksheet['G4']
                phone_sans_frais.value = 'Sans frais (1 800 263-2572, poste 56221)'

                worksheet.merge_cells('AA4:AK4') # Changed from AA5:AK5
                legend_d = worksheet['AA4']
                legend_d.value = 'D9h ou D17h : la garde Débute à 9h ou 17h'
                cyan_fill = PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid')
                legend_d.fill = cyan_fill
                
                # Row 5
                worksheet.merge_cells('AA5:AK5') # Changed from AA5:AK5
                legend_f = worksheet['AA5']
                legend_f.value = 'F9h ou F17h : la garde Finit à 9h ou 17h'
                legend_f.fill = cyan_fill
                # --- HEADER FORMATTING (MODIFICATIONS END) ---
                
                # Row 6: Empty row for spacing

                # Set row heights
                worksheet.row_dimensions[1].height = 30
                for row_idx in range(2, 7):
                    worksheet.row_dimensions[row_idx].height = 20
                worksheet.row_dimensions[6].height = 10
                for row_idx in range(8, 11):
                    worksheet.row_dimensions[row_idx].height = 15
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 3
                worksheet.column_dimensions['B'].width = 15
                for col_idx in range(3, len(date_range) + 10):  # just enough columns for the dates
                    worksheet.column_dimensions[get_column_letter(col_idx)].width = 3.7
                
                # --- FORMATTING STYLES ---
                day_number_fill = PatternFill(start_color='2F2F2F', end_color='2F2F2F', fill_type='solid')
                day_symbol_fill = PatternFill(start_color='6B6969', end_color='6B6969', fill_type='solid')
                white_bold_font = Font(bold=True, color='FFFFFF')
                weekend_fill = PatternFill(start_color='8FD6F3', end_color='8FD6F3', fill_type='solid')
                team_fill = PatternFill(start_color='D9F0FB', end_color='D9F0FB', fill_type='solid')
                bold_font = Font(bold=True)
                center_alignment_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
                thin_side = Side(border_style="thin", color="000000")
                thick_side = Side(border_style="thick", color="000000")
                grey_thin_side = Side(border_style="thin", color="BFBFBF")
                thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                grey_border_for_headers = Border(left=grey_thin_side, right=grey_thin_side, top=grey_thin_side, bottom=grey_thin_side)
                
                # --- SCHEDULE HEADER FORMATTING ---
                worksheet.merge_cells('B8:B9')
                sage_femme_header = worksheet['B8']
                sage_femme_header.value = 'Sage-femme'
                sage_femme_header.fill = day_number_fill
                sage_femme_header.font = white_bold_font
                sage_femme_header.alignment = center_alignment_wrap
                sage_femme_header.border = grey_border_for_headers
                
                for col_idx, date in enumerate(date_range):
                    col_letter = get_column_letter(col_idx + 3)
                    
                    day_num_cell = worksheet[f"{col_letter}8"]
                    day_num_cell.value = date.day
                    day_num_cell.fill = day_number_fill
                    day_num_cell.font = white_bold_font
                    day_num_cell.alignment = center_alignment_wrap
                    day_num_cell.border = grey_border_for_headers
                    
                    weekday_index = (date.weekday() + 1) % 7
                    day_symbol = ['D', 'L', 'M', 'Me', 'J', 'V', 'S'][weekday_index]
                    day_symbol_cell = worksheet[f"{col_letter}9"]
                    day_symbol_cell.value = day_symbol
                    day_symbol_cell.fill = day_symbol_fill
                    day_symbol_cell.font = white_bold_font
                    day_symbol_cell.alignment = center_alignment_wrap
                    day_symbol_cell.border = grey_border_for_headers
                
                month_font = Font(bold=True, size=16)
                center_alignment = Alignment(horizontal='center', vertical='center')
                months_data = {}
                for col_idx, date in enumerate(date_range):
                    french_months = {1: 'JANVIER', 2: 'FÉVRIER', 3: 'MARS', 4: 'AVRIL', 5: 'MAI', 6: 'JUIN', 7: 'JUILLET', 8: 'AOÛT', 9: 'SEPTEMBRE', 10: 'OCTOBRE', 11: 'NOVEMBRE', 12: 'DÉCEMBRE'}
                    month_year = french_months[date.month]
                    if month_year not in months_data:
                        months_data[month_year] = {'start_col': col_idx + 3}
                    months_data[month_year]['end_col'] = col_idx + 3

                for month_name, month_info in months_data.items():
                    start_col_letter = get_column_letter(month_info['start_col'])
                    end_col_letter = get_column_letter(month_info['end_col'])
                    if start_col_letter != end_col_letter:
                        worksheet.merge_cells(f"{start_col_letter}7:{end_col_letter}7")
                    
                    month_cell = worksheet[f"{start_col_letter}7"]
                    month_cell.value = month_name
                    month_cell.font = month_font
                    month_cell.alignment = center_alignment
                
                # --- MAIN TABLE FORMATTING ---
                max_col = len(df.columns) + 1
                current_row = 10  # Start at row 10 to match where data is written (startrow=9 means row 10 in Excel)
                for team_index, team in enumerate(teams_data):
                    # Get team members - use the member data directly from the frontend
                    team_members = team.get('members', [])
                    team_members_with_schedule_data = []
                    
                    for member in team_members:
                        member_name = member.get('name', '')
                        if not member_name:
                            continue
                        
                        # Check if we have schedule data for this member
                        has_schedule_data = member_name in worker_schedule_data
                        if not has_schedule_data:
                            # Try partial name matching
                            member_name_parts = member_name.lower().split(' ')
                            for worker_name in worker_schedule_data.keys():
                                worker_name_parts = worker_name.lower().split(' ')
                                if any(part in worker_name_parts for part in member_name_parts if len(part) > 2):
                                    has_schedule_data = True
                                    break
                        
                        if has_schedule_data or True:  # Include all members even if no schedule data
                            team_members_with_schedule_data.append(member_name)
                    
                    num_members = len(team_members_with_schedule_data)

                    if num_members == 0:
                        continue
                    
                    start_row = current_row
                    end_row = current_row + num_members - 1
                    is_team_to_color = (team_index % 2 == 0)
                    
                    for member_offset in range(num_members):
                        row_idx = current_row + member_offset
                        worksheet.row_dimensions[row_idx].height = 26.25 # Set row height
                        for col_idx in range(2, max_col + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.alignment = center_alignment_wrap
                            cell.border = thin_border
                            
                            # Apply font styles
                            if col_idx == 2: # Name column
                                cell.font = franklin_gothic_book_font_11_bold
                            else: # Schedule data columns
                                cell.font = schedule_font
                            
                            is_weekend = False
                            if col_idx > 2:
                                date_idx = col_idx - 3
                                if date_idx < len(date_range):
                                    date = date_range[date_idx]
                                    if (date.weekday() + 1) % 7 in [0, 6]:
                                        is_weekend = True
                            
                            if is_weekend:
                                cell.fill = weekend_fill
                            elif is_team_to_color:
                                cell.fill = team_fill
                    
                    # Apply borders to the team block
                    for r in range(start_row, end_row + 1):
                        for c in range(2, max_col + 1):
                            cell = worksheet.cell(row=r, column=c)
                            top_b = thick_side if r == start_row else thin_side
                            bottom_b = thick_side if r == end_row else thin_side
                            left_b = thick_side if c == 2 else thin_side
                            right_b = thick_side if c == max_col else thin_side
                            
                            if c == 2: right_b = thick_side
                            if c == 3: left_b = thick_side
                            cell.border = Border(left=left_b, right=right_b, top=top_b, bottom=bottom_b)
                    current_row += num_members

                # --- FOOTER AND LEGEND ---
                bottom_day_num_row = current_row  # Use current_row which is now 1 row below the last data row
                for col_idx, date in enumerate(date_range):
                    col_letter = get_column_letter(col_idx + 3)
                    day_num_cell_bottom = worksheet[f"{col_letter}{bottom_day_num_row}"]
                    day_num_cell_bottom.value = date.day
                    day_num_cell_bottom.fill = day_number_fill
                    day_num_cell_bottom.font = white_bold_font
                    day_num_cell_bottom.alignment = center_alignment_wrap
                
                # --- CONTACT INFO FOOTER ---
                conn = get_db_connection()
                workers_with_phones = conn.execute('SELECT name, phone_number FROM workers').fetchall()
                conn.close()
                
                phone_data = {worker['name'].strip().upper(): worker['phone_number'] for worker in workers_with_phones if worker['phone_number']}
                
                footer_start_row = bottom_day_num_row + 2
                contact_info_font = Font(bold=True)
                contact_info_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                start_col = 3
                cols_per_contact = 3
                current_col = start_col
                
                has_members = False
                all_contact_cols = []
                total_members_count = 0
                current_member_index = 0

                # First, count total members across all teams using frontend data
                for team in teams_data:
                    team_members = team.get('members', [])
                    for member in team_members:
                        member_name = member.get('name', '')
                        if member_name:
                            total_members_count += 1

                for team_index, team in enumerate(teams_data):
                    # Get team members - use the member data directly from the frontend
                    team_members = team.get('members', [])
                    team_member_names = [member.get('name', '') for member in team_members if member.get('name', '')]
                    
                    if not team_member_names:
                        continue
                    
                    has_members = True
                    is_team_to_color = (team_index % 2 == 0)
                    is_first_team = (team_index == 0)
                    is_last_team = (team_index == len([t for t in teams_data if t.get('members')]) - 1)
                    
                    for member_index, member_name in enumerate(team_member_names):
                        is_first_member_of_team = (member_index == 0)
                        is_last_member_of_team = (member_index == len(team_member_names) - 1)
                        
                        contact_fill = team_fill if is_team_to_color else PatternFill(fill_type=None)

                        name_start_col_letter = get_column_letter(current_col)
                        name_end_col_letter = get_column_letter(current_col + cols_per_contact - 1)
                        worksheet.merge_cells(f"{name_start_col_letter}{footer_start_row}:{name_end_col_letter}{footer_start_row}")
                        
                        name_cell = worksheet.cell(row=footer_start_row, column=current_col)
                        name_cell.value = member_name
                        name_cell.font = contact_info_font
                        name_cell.alignment = contact_info_alignment
                        name_cell.fill = contact_fill
                        
                        phone_start_col_letter = get_column_letter(current_col)
                        phone_end_col_letter = get_column_letter(current_col + cols_per_contact - 1)
                        worksheet.merge_cells(f"{phone_start_col_letter}{footer_start_row + 1}:{phone_end_col_letter}{footer_start_row + 1}")
                        
                        phone_number = phone_data.get(member_name.upper(), '')
                        phone_cell = worksheet.cell(row=footer_start_row + 1, column=current_col)
                        phone_cell.value = phone_number
                        phone_cell.alignment = contact_info_alignment
                        phone_cell.fill = contact_fill

                        is_first_member_of_all = (current_member_index == 0)
                        is_last_member_of_all = (current_member_index == total_members_count - 1)

                        for r in range(footer_start_row, footer_start_row + 2):
                            for c_offset in range(cols_per_contact):
                                col_to_border = current_col + c_offset
                                cell = worksheet.cell(row=r, column=col_to_border)
                                
                                left_b, right_b, top_b, bottom_b = cell.border.left, cell.border.right, cell.border.top, cell.border.bottom

                                # Left border: thick if first member of team or first overall
                                if c_offset == 0:
                                    if is_first_member_of_team or is_first_member_of_all:
                                        left_b = thick_side
                                    else:
                                        left_b = thin_side
                                
                                # Right border: thick if last member of team or last overall
                                if c_offset == cols_per_contact - 1:
                                    if is_last_member_of_team or is_last_member_of_all:
                                        right_b = thick_side
                                    else:
                                        right_b = thin_side
                                
                                cell.border = Border(left=left_b, right=right_b, top=top_b, bottom=bottom_b)

                        all_contact_cols.extend(range(current_col, current_col + cols_per_contact))
                        current_col += cols_per_contact
                        current_member_index += 1

                if has_members:
                    worksheet.row_dimensions[footer_start_row].height = 30
                    worksheet.row_dimensions[footer_start_row + 1].height = 30
                    
                    contact_start_col = 3
                    contact_end_col = max(all_contact_cols) if all_contact_cols else start_col

                    for r in range(footer_start_row, footer_start_row + 2):
                        for c in range(contact_start_col, contact_end_col + 1):
                            cell = worksheet.cell(row=r, column=c)
                            top_border = thick_side
                            bottom_border = thick_side
                            if r == footer_start_row: bottom_border = thin_side
                            if r == footer_start_row + 1: top_border = thin_side
                            
                            cell.border = Border(left=cell.border.left, right=cell.border.right, top=top_border, bottom=bottom_border)

        except Exception as e:
            print(f"Error exporting team schedules to Excel: {e}")
            traceback.print_exc()
            user_error = f"Export failed: {str(e)}"
            if "Permission denied" in str(e):
                user_error = "Unable to save the Excel file. Please close any open Excel files with the same name and try again."
            return jsonify({"error": user_error}), 500

        return jsonify({
            "message": "Horaires des équipes exportés avec succès",
            "filename": filename,
        })

    except Exception as e:
        print(f"Error in export_team_schedules: {e}")
        traceback.print_exc()
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


# NEW ENDPOINTS for employee management
@app.route('/api/workers', methods=['POST'])
def add_worker():
    try:
        data = request.get_json()
        name = data.get('name')
        status = data.get('status', 'employee')
        employee_number = data.get('employee_number')
        phone_number = data.get('phone_number')
        
        if not name:
            return jsonify({"error": "Name is required"}), 400
        
        if not employee_number:
            return jsonify({"error": "Employee number is required"}), 400
        
        conn = get_db_connection()
        
        # Check if employee_number column exists
        cursor = conn.cursor()
        cursor.execute('PRAGMA table_info(workers)')
        columns = [row[1] for row in cursor.fetchall()]
        
        # Check for duplicate employee number if column exists
        if 'employee_number' in columns:
            existing = conn.execute('SELECT id FROM workers WHERE employee_number = ?', (employee_number,)).fetchone()
            if existing:
                conn.close()
                return jsonify({"error": "Employee number already exists"}), 400
        
        if 'employee_number' in columns:
            cursor = conn.execute('INSERT INTO workers (name, status, employee_number, phone_number) VALUES (?, ?, ?, ?)', 
                                 (name, status, employee_number, phone_number))
        else:
            # Fallback for databases without employee_number column
            cursor = conn.execute('INSERT INTO workers (name, status) VALUES (?, ?)', 
                                 (name, status))
        
        worker_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({"id": worker_id, "name": name, "status": status, "employee_number": employee_number, "phone_number": phone_number}), 201
    except Exception as e:
        print(f"Error adding worker: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/workers/<int:worker_id>', methods=['PUT'])
def update_worker(worker_id):
    try:
        data = request.get_json()
        name = data.get('name')
        status = data.get('status')
        employee_number = data.get('employee_number')
        phone_number = data.get('phone_number')
        
        if not name:
            return jsonify({"error": "Name is required"}), 400
        
        if not employee_number:
            return jsonify({"error": "Employee number is required"}), 400
        
        conn = get_db_connection()
        
        # Check if employee_number column exists
        cursor = conn.cursor()
        cursor.execute('PRAGMA table_info(workers)')
        columns = [row[1] for row in cursor.fetchall()]
        
        # Check for duplicate employee number (excluding current worker) if column exists
        if 'employee_number' in columns:
            existing = conn.execute('SELECT id FROM workers WHERE employee_number = ? AND id != ?', (employee_number, worker_id)).fetchone()
            if existing:
                conn.close()
                return jsonify({"error": "Employee number already exists"}), 400
        
        if 'employee_number' in columns:
            conn.execute('UPDATE workers SET name = ?, status = ?, employee_number = ?, phone_number = ? WHERE id = ?', 
                         (name, status, employee_number, phone_number, worker_id))
        else:
            # Fallback for databases without employee_number column
            conn.execute('UPDATE workers SET name = ?, status = ? WHERE id = ?', 
                         (name, status, worker_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({"id": worker_id, "name": name, "status": status, "employee_number": employee_number, "phone_number": phone_number})
    except Exception as e:
        print(f"Error updating worker: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/workers/<int:worker_id>', methods=['DELETE'])
def delete_worker(worker_id):
    try:
        conn = get_db_connection()
        conn.execute('DELETE FROM workers WHERE id = ?', (worker_id,))
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Worker deleted successfully"})
    except Exception as e:
        print(f"Error deleting worker: {e}")
        return jsonify({"error": str(e)}), 500

# NEW ENDPOINTS for employee schedule management

@app.route('/api/schedules/weekly', methods=['POST'])
def add_weekly_schedule():
    """Add a weekly schedule for an employee over a date range"""
    conn = None
    try:
        data = request.get_json()
        employee_id = data.get('employeeId')
        start_date = data.get('startDate')
        end_date = data.get('endDate')
        weekly_schedule = data.get('weeklySchedule', [])
        
        # New: Handle indisponibilité data
        unavailability_type = data.get('unavailabilityType')
        unavailability_start = data.get('unavailabilityStartDate')
        unavailability_end = data.get('unavailabilityEndDate')
        
        if not employee_id or not start_date or not end_date:
            return jsonify({"error": "Employee ID, start date, and end date are required"}), 400
        
        # Parse dates
        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
        
        # Parse unavailability dates if provided
        unavailability_start_dt = None
        unavailability_end_dt = None
        if unavailability_type and unavailability_start and unavailability_end:
            unavailability_start_dt = datetime.strptime(unavailability_start, '%Y-%m-%d')
            unavailability_end_dt = datetime.strptime(unavailability_end, '%Y-%m-%d')
            print(f"Unavailability period: {unavailability_start_dt.strftime('%Y-%m-%d')} to {unavailability_end_dt.strftime('%Y-%m-%d')} ({unavailability_type})")
        
        conn = get_db_connection()
        
        # Ensure table exists with proper schema
        migrate_database()
        
        # Generate dates for the range and insert schedule
        current_date = start_dt
        while current_date <= end_dt:
            day_of_week = current_date.weekday()  # Monday = 0, Sunday = 6
            # Convert to our format where Sunday = 0
            day_index = (day_of_week + 1) % 7
            
            # Check if current date falls within unavailability period
            is_unavailable = False
            if (unavailability_start_dt and unavailability_end_dt and 
                unavailability_start_dt <= current_date <= unavailability_end_dt):
                is_unavailable = True
            
            # Delete existing schedules for this day first
            conn.execute('''
                DELETE FROM employee_schedules 
                WHERE employee_id = ? AND date = ?
            ''', (employee_id, current_date.strftime('%Y-%m-%d')))
            
            if is_unavailable:
                # Insert unavailability record (M or V)
                print(f"Inserting unavailability record for {current_date.strftime('%Y-%m-%d')}: {unavailability_type}")
                conn.execute('''
                    INSERT INTO employee_schedules 
                    (employee_id, date, day_of_week, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_number, hours_worked)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (employee_id, current_date.strftime('%Y-%m-%d'), day_index, 
                      None, None, 'unavailable', unavailability_type, False, 1, 0))
            else:
                # Find matching schedules for this day (could be multiple shifts)
                day_schedules = [s for s in weekly_schedule if s['day'] == day_index]
                
                if day_schedules:
                    # Insert each shift for this day
                    for shift_num, day_schedule in enumerate(day_schedules, 1):
                        hours_worked = calculate_hours_worked(
                            day_schedule['startTime'], 
                            day_schedule['endTime'], 
                            day_schedule.get('isNightShift', False)
                        )
                        
                        conn.execute('''
                            INSERT INTO employee_schedules 
                            (employee_id, date, day_of_week, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_number, hours_worked)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        ''', (employee_id, current_date.strftime('%Y-%m-%d'), day_index, 
                              day_schedule['startTime'], day_schedule['endTime'], 'work',
                              day_schedule.get('activityType', 'X'),
                              day_schedule.get('isNightShift', False),
                              shift_num, hours_worked))
                        
                        # Note: F\n9h entries are now generated dynamically during display/export
                        # No need to store them in the database
            
            current_date = current_date + timedelta(days=1)
        
        conn.commit()
        return jsonify({"message": "Weekly schedule added successfully"})
        
    except Exception as e:
        print(f"Error adding weekly schedule: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/schedules/day', methods=['PUT'])
def update_day_schedule():
    """Update a specific day schedule for an employee"""
    conn = None
    try:
        data = request.get_json()
        employee_id = data.get('employeeId')
        date = data.get('date')
        start_time = data.get('startTime')
        end_time = data.get('endTime')
        activity_type = data.get('activityType', 'X')
        is_night_shift = data.get('isNightShift', False)
        
        if not employee_id or not date or not start_time or not end_time:
            return jsonify({"error": "Employee ID, date, start time, and end time are required"}), 400
        
        conn = get_db_connection()
        
        # Calculate hours worked
        hours_worked = calculate_hours_worked(start_time, end_time, is_night_shift)
        
        # Get day of week
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        day_of_week = (date_obj.weekday() + 1) % 7  # Convert to Sunday = 0
        
        # Delete existing schedules for this day first
        conn.execute('''
            DELETE FROM employee_schedules 
            WHERE employee_id = ? AND date = ?
        ''', (employee_id, date))
        
        # Insert updated schedule
        conn.execute('''
            INSERT INTO employee_schedules 
            (employee_id, date, day_of_week, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_number, hours_worked, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (employee_id, date, day_of_week, start_time, end_time, 'work', activity_type, is_night_shift, 1, hours_worked))
        
        # Handle night shift end marking if applicable
        # Note: F\n9h entries are now generated dynamically during display/export
        # No need to store them in the database
        
        conn.commit()
        return jsonify({"message": "Day schedule updated successfully"})
        
    except Exception as e:
        print(f"Error updating day schedule: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/schedules/day', methods=['DELETE'])
def delete_day_schedule():
    """Delete a specific day schedule for an employee"""
    conn = None
    try:
        data = request.get_json()
        employee_id = data.get('employeeId')
        date = data.get('date')
        
        if not employee_id or not date:
            return jsonify({"error": "Employee ID and date are required"}), 400
        
        conn = get_db_connection()
        
        # Check if this is a night shift to handle cleanup
        result = conn.execute('''
            SELECT is_night_shift, end_time FROM employee_schedules 
            WHERE employee_id = ? AND date = ? AND schedule_type = 'work'
        ''', (employee_id, date)).fetchone()
        
        is_night_shift = False
        end_time = None
        if result:
            is_night_shift = result[0]
            end_time = result[1]
        
        # Delete the main schedule entry
        conn.execute('''
            DELETE FROM employee_schedules 
            WHERE employee_id = ? AND date = ?
        ''', (employee_id, date))
        
        # If it was a night shift, also delete the next day's end marker
        if is_night_shift and end_time:
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            next_date = date_obj + timedelta(days=1)
            conn.execute('''
                DELETE FROM employee_schedules 
                WHERE employee_id = ? AND date = ? AND schedule_type = 'end_of_night_shift'
            ''', (employee_id, next_date.strftime('%Y-%m-%d')))
        
        conn.commit()
        return jsonify({"message": "Day schedule deleted successfully"})
        
    except Exception as e:
        print(f"Error deleting day schedule: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/workers/<int:worker_id>/calendar', methods=['GET'])
def get_worker_calendar(worker_id):
    """Get calendar data for a specific worker"""
    try:
        year = request.args.get('year', default=datetime.now().year, type=int)
        month = request.args.get('month', default=datetime.now().month, type=int)
        
        conn = get_db_connection()
        
        # Create table if it doesn't exist (updated schema)
        conn.execute('''
            CREATE TABLE IF NOT EXISTS employee_schedules (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                employee_id INTEGER,
                date DATE,
                day_of_week INTEGER,
                start_time TEXT,
                end_time TEXT,
                schedule_type TEXT DEFAULT 'work',
                activity_type TEXT DEFAULT 'X',
                is_night_shift BOOLEAN DEFAULT 0,
                shift_number INTEGER DEFAULT 1,
                hours_worked REAL DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (employee_id) REFERENCES workers (id)
            )
        ''')
        
        # Ensure new columns exist for existing databases
        try:
            conn.execute('ALTER TABLE employee_schedules ADD COLUMN activity_type TEXT DEFAULT "X"')
        except sqlite3.OperationalError:
            pass  # Column already exists
            
        try:
            conn.execute('ALTER TABLE employee_schedules ADD COLUMN shift_number INTEGER DEFAULT 1')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        try:
            conn.execute('ALTER TABLE employee_schedules ADD COLUMN hours_worked REAL DEFAULT 0')
        except sqlite3.OperationalError:
            pass  # Column already exists
        
        # Get schedules for the specified month
        start_date = f"{year}-{month:02d}-01"
        if month == 12:
            end_date = f"{year + 1}-01-01"
        else:
            end_date = f"{year}-{month + 1:02d}-01"

        schedules = conn.execute('''
            SELECT date, start_time, end_time, schedule_type, activity_type, is_night_shift, hours_worked, shift_number
            FROM employee_schedules
            WHERE employee_id = ? AND date >= ? AND date < ?
            ORDER BY date, shift_number
        ''', (worker_id, start_date, end_date)).fetchall()

        conn.close()

        # Convert to dictionary format, handling multiple shifts per day
        calendar_data = {}
        for schedule in schedules:
            date_str = schedule['date']
            
            if date_str not in calendar_data:
                calendar_data[date_str] = []
            
            calendar_data[date_str].append({
                'start_time': schedule['start_time'],
                'end_time': schedule['end_time'],
                'schedule_type': schedule['schedule_type'],
                'activity_type': schedule['activity_type'],
                'is_night_shift': bool(schedule['is_night_shift']),
                'hours_worked': schedule['hours_worked'],
                'shift_number': schedule['shift_number']
            })
        
        # For single shifts, maintain backwards compatibility by returning the shift directly
        simplified_calendar_data = {}
        for date_str, shifts in calendar_data.items():
            if len(shifts) == 1:
                simplified_calendar_data[date_str] = shifts[0]
            else:
                simplified_calendar_data[date_str] = {
                    'schedule_type': 'multiple_shifts',
                    'shifts': shifts,
                    'total_hours': sum(shift['hours_worked'] or 0 for shift in shifts if shift['schedule_type'] == 'work'),
                    'activity_types': list(set(shift['activity_type'] for shift in shifts if shift['schedule_type'] == 'work'))
                }

        return jsonify(simplified_calendar_data)
        
    except Exception as e:
        print(f"Error getting worker calendar: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/workers/<int:worker_id>/schedule/day', methods=['GET'])
def get_day_schedule_details(worker_id):
    """Get detailed schedule information for a specific day"""
    conn = None
    try:
        date = request.args.get('date')
        if not date:
            return jsonify({"error": "Date parameter is required"}), 400
        
        conn = get_db_connection()
        
        # Get all shifts for this worker on this specific date
        cursor = conn.execute('''
            SELECT id, employee_id, date, start_time, end_time, 
                   schedule_type, activity_type, is_night_shift, 
                   shift_number, hours_worked
            FROM employee_schedules 
            WHERE employee_id = ? AND date = ?
            ORDER BY shift_number
        ''', (worker_id, date))
        
        shifts = cursor.fetchall()
        
        if not shifts:
            return jsonify([]), 200
        
        # Convert to list of dictionaries
        result = []
        for shift in shifts:
            shift_dict = {
                'id': shift['id'],
                'employee_id': shift['employee_id'],
                'date': shift['date'],
                'start_time': shift['start_time'],
                'end_time': shift['end_time'],
                'schedule_type': shift['schedule_type'],
                'activity_type': shift['activity_type'],
                'is_night_shift': bool(shift['is_night_shift']),
                'shift_number': shift['shift_number'],
                'hours_worked': shift['hours_worked']
            }
            result.append(shift_dict)
        
        return jsonify(result)
        
    except Exception as e:
        print(f"Error getting day schedule details: {e}")
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

@app.route('/api/workers/<int:worker_id>/schedule/day', methods=['POST'])
def save_day_schedule(worker_id):
    """Save or update a day schedule for a worker (supports multiple shifts)"""
    conn = None
    try:
        data = request.get_json()
        date = data.get('date')
        shifts = data.get('shifts', [])
        
        # Handle single shift for backwards compatibility
        if not shifts and data.get('schedule_type'):
            shifts = [{
                'schedule_type': data.get('schedule_type', 'work'),
                'start_time': data.get('start_time'),
                'end_time': data.get('end_time'),
                'activity_type': data.get('activity_type', 'X'),
                'is_night_shift': data.get('isNightShift', False)
            }]
        
        if not date:
            return jsonify({"error": "Date is required"}), 400
        
        # Parse date to get day of week
        date_obj = datetime.strptime(date, '%Y-%m-%d')
        day_of_week = (date_obj.weekday() + 1) % 7  # Convert to Sunday = 0 format
        
        conn = get_db_connection()
        
        # Ensure table exists with proper schema
        migrate_database()
        
        # Delete existing schedules for this day
        conn.execute('''
            DELETE FROM employee_schedules 
            WHERE employee_id = ? AND date = ?
        ''', (worker_id, date))
        
        # Insert new shifts
        total_hours = 0
        for shift_num, shift_data in enumerate(shifts, 1):
            schedule_type = shift_data.get('schedule_type', 'work')
            start_time = shift_data.get('start_time')
            end_time = shift_data.get('end_time')
            activity_type = shift_data.get('activity_type', 'X')
            is_night_shift = shift_data.get('isNightShift', False)
            
            # Calculate hours worked for work schedules
            hours_worked = 0
            if schedule_type == 'work' and start_time and end_time:
                hours_worked = calculate_hours_worked(start_time, end_time, is_night_shift)
                total_hours += hours_worked
            
            # Insert the schedule (simple INSERT since we already deleted existing ones)
            conn.execute('''
                INSERT INTO employee_schedules 
                (employee_id, date, day_of_week, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_number, hours_worked, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (worker_id, date, day_of_week, start_time, end_time, schedule_type, activity_type, is_night_shift, shift_num, hours_worked))
            
            # Note: F\n9h entries are now generated dynamically during display/export
            # No need to store them in the database
        
        conn.commit()
        
        return jsonify({
            "message": "Day schedule saved successfully",
            "total_hours_worked": total_hours,
            "shifts_count": len(shifts),
            "shifts": shifts
        })
        
    except Exception as e:
        print(f"Error saving day schedule: {e}")
        import traceback
        traceback.print_exc()
        if conn:
            conn.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        if conn:
            conn.close()

def calculate_hours_worked(start_time, end_time, is_night_shift=False):
    """Calculate hours worked between start and end time"""
    try:
        from datetime import datetime, time
        
        # Parse time strings
        start = datetime.strptime(start_time, '%H:%M').time()
        end = datetime.strptime(end_time, '%H:%M').time()
        
        # Convert to datetime objects for calculation
        start_dt = datetime.combine(datetime.today(), start)
        end_dt = datetime.combine(datetime.today(), end)
        
        # For night shifts, if end time is before start time, it means next day
        if is_night_shift and end < start:
            end_dt = end_dt.replace(day=end_dt.day + 1)
        elif not is_night_shift and end <= start:
            # For non-night shifts, if end is before or equal to start, assume next day
            end_dt = end_dt.replace(day=end_dt.day + 1)
        
        # Calculate difference in hours
        duration = end_dt - start_dt
        hours = duration.total_seconds() / 3600
        
        return round(hours, 2)
        
    except Exception as e:
        print(f"Error calculating hours: {e}")
        return 0

@app.route('/api/workers/<int:worker_id>/weekly-summary', methods=['GET'])
def get_worker_weekly_summary(worker_id):
    """Get weekly summary statistics for a worker"""
    try:
        date_str = request.args.get('date', default=datetime.today().strftime('%Y-%m-%d'))
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        
        # Find the Sunday of that week (start of week)
        day_of_week = date_obj.weekday()  # Monday = 0, Sunday = 6
        days_to_subtract = (day_of_week + 1) % 7  # Convert to Sunday = 0 format
        week_start = date_obj - timedelta(days=days_to_subtract)
        week_end = week_start + timedelta(days=6)
        
        conn = get_db_connection()
        
        # Get schedules for the week
        schedules = conn.execute('''
            SELECT date, start_time, end_time, schedule_type, activity_type, is_night_shift, hours_worked
            FROM employee_schedules
            WHERE employee_id = ? AND date >= ? AND date <= ?
            ORDER BY date
        ''', (worker_id, week_start.strftime('%Y-%m-%d'), week_end.strftime('%Y-%m-%d'))).fetchall()
        
        conn.close()
        
        # Calculate summary statistics
        total_hours = 0
        work_days = 0
        night_shifts = 0
        activity_breakdown = {}
        
        for schedule in schedules:
            if schedule['schedule_type'] == 'work':
                work_days += 1
                total_hours += schedule['hours_worked'] or 0
                
                if schedule['is_night_shift']:
                    night_shifts += 1
                
                activity = schedule['activity_type'] or 'work'
                activity_breakdown[activity] = activity_breakdown.get(activity, 0) + 1
        
        return jsonify({
            "week_start": week_start.strftime('%Y-%m-%d'),
            "week_end": week_end.strftime('%Y-%m-%d'),
            "total_hours": round(total_hours, 2),
            "work_days": work_days,
            "night_shifts": night_shifts,
            "activity_breakdown": activity_breakdown,
            "overtime_hours": max(0, total_hours - 40) if total_hours > 40 else 0
        })
        
    except Exception as e:
        print(f"Error getting weekly summary: {e}")
        return jsonify({"error": str(e)}), 500

from datetime import timedelta

# NEW ENDPOINTS for reports management

@app.route('/api/reports/archive', methods=['GET'])
def get_archive_files():
    """Get list of archive files"""
    try:
        archive_dir = get_archive_directory()
        files_list = []
        
        for filename in os.listdir(archive_dir):
            # Skip temporary Excel files (those starting with ~$)
            if filename.startswith('~$'):
                continue
                
            if filename.endswith('.xlsx') or filename.endswith('.xls'):
                file_path = os.path.join(archive_dir, filename)
                file_stats = os.stat(file_path)
                
                files_list.append({
                    'filename': filename,
                    'size': file_stats.st_size,
                    'created_date': datetime.fromtimestamp(file_stats.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_date': datetime.fromtimestamp(file_stats.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
        
        # Sort by modification date (newest first)
        files_list.sort(key=lambda x: x['modified_date'], reverse=True)
        
        return jsonify(files_list)
            
    except Exception as e:
        print(f"Error getting archive files: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/reports/archive/<filename>', methods=['PUT'])
def update_archive_file(filename):
    """Update (regenerate) an archive file"""
    try:
        # Reject temporary Excel files
        if filename.startswith('~$'):
            return jsonify({"error": "Cannot update temporary Excel files"}), 400
        
        # Extract date from filename (support multiple formats)
        date_str = None
        
        # Format 1: weekly_report_YYYY-MM-DD.xlsx
        if filename.startswith('weekly_report_') and filename.endswith('.xlsx'):
            date_str = filename.replace('weekly_report_', '').replace('.xlsx', '')
        # Format 2: rapport_hebdomadaire_YYYY-MM-DD.xlsx (case insensitive)
        elif filename.lower().startswith('rapport_hebdomadaire_') and filename.endswith('.xlsx'):
            # Extract the date part after 'rapport_hebdomadaire_'
            date_str = filename[len('rapport_hebdomadaire_'):-5]  # Remove prefix and .xlsx
        # Format 3: Horaires_Equipes_YYYY-MM-DD_au_YYYY-MM-DD.xlsx
        elif filename.startswith('Horaires_Equipes_') and filename.endswith('.xlsx'):
            # This is a team schedule file - handle regeneration
            print(f"Regenerating team schedule file: {filename}")
            
            # Extract date range from filename
            # Format: Horaires_Equipes_2025-08-20_au_2025-09-30.xlsx
            filename_without_ext = filename.replace('.xlsx', '')
            parts = filename_without_ext.split('_')
            
            if len(parts) >= 5 and parts[3] == 'au':
                start_date_str = parts[2]
                end_date_str = parts[4]
                
                try:
                    # Validate dates
                    start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
                    end_date = datetime.strptime(end_date_str, '%Y-%m-%d')
                    
                    # Get teams data
                    conn = get_db_connection()
                    teams_query = '''
                        SELECT t.id, t.name, t.description, t.color, t.created_at,
                               COUNT(DISTINCT tm.employee_id) as member_count
                        FROM teams t
                        LEFT JOIN team_members tm ON t.id = tm.team_id
                        GROUP BY t.id, t.name, t.description, t.color, t.created_at
                        HAVING member_count > 0
                        ORDER BY t.created_at DESC
                    '''
                    teams = conn.execute(teams_query).fetchall()
                    
                    # Build teams data with members
                    teams_data = []
                    for team in teams:
                        members_query = '''
                            SELECT tm.employee_id as id, w.name, w.employee_number
                            FROM team_members tm
                            JOIN workers w ON tm.employee_id = w.id
                            WHERE tm.team_id = ?
                            ORDER BY w.name
                        '''
                        members = conn.execute(members_query, (team['id'],)).fetchall()
                        team_data = {
                            'id': team['id'],
                            'name': team['name'],
                            'description': team['description'],
                            'color': team['color'],
                            'members': [{'id': m['id'], 'name': m['name'], 'employee_number': m['employee_number']} for m in members]
                        }
                        if team_data['members']:  # Only include teams with members
                            teams_data.append(team_data)
                    
                    conn.close()
                    
                    # Call the team schedule export function directly
                    from flask import request
                    import json
                    
                    # Simulate the request data that the export function expects
                    export_data = {
                        'start_date': start_date_str,
                        'end_date': end_date_str,
                        'teams': teams_data
                    }
                    
                    # Save original request and create mock request
                    original_request = request
                    
                    # Create a mock request for the export function
                    class MockRequest:
                        def get_json(self):
                            return export_data
                    
                    # Temporarily replace request
                    import sys
                    current_module = sys.modules[__name__]
                    setattr(current_module, 'request', MockRequest())
                    
                    try:
                        # Call the export function
                        result = export_team_schedules()
                        
                        # Check if export was successful
                        if hasattr(result, 'status_code') and result.status_code != 200:
                            return result
                        
                        return jsonify({
                            "message": f"Team schedule file regenerated successfully: {filename}",
                            "filename": filename
                        })
                        
                    finally:
                        # Restore original request
                        setattr(current_module, 'request', original_request)
                        
                except ValueError as e:
                    return jsonify({"error": f"Invalid date format in filename: {str(e)}"}), 400
                except Exception as e:
                    print(f"Error regenerating team schedule: {e}")
                    return jsonify({"error": f"Failed to regenerate team schedule: {str(e)}"}), 500
            else:
                return jsonify({"error": "Invalid team schedule filename format"}), 400
        
        if date_str:
            try:
                datetime.strptime(date_str, '%Y-%m-%d')  # Validate date format
            except ValueError:
                return jsonify({"error": "Invalid filename format"}), 400
        else:
            return jsonify({"error": "Invalid filename format"}), 400
        
        # Change to the core directory
        original_cwd = os.getcwd()
        core_dir = os.path.join(os.path.dirname(__file__), 'core')
        os.chdir(core_dir)
        
        try:
            week_start = datetime.strptime(date_str, '%Y-%m-%d')
            
            # Get the dataframe for the week
            df = processor.getDataframe(week_start)
            
            # If df is None or empty, create empty dataframe
            if df is None or df.empty:
                print("DataFrame is empty, creating empty dataframe for Excel export")
                # Create empty dataframe with proper structure
                df = pd.DataFrame(columns=["Name", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "TOTAL"])
            
            # Create the file path using standardized archive directory
            archive_dir = os.path.join('data', 'archive')
            os.makedirs(archive_dir, exist_ok=True)
            archive_path = os.path.join(archive_dir, filename)
            
            # Create Excel file
            import xlsxwriter
            workbook = xlsxwriter.Workbook(archive_path)
            worksheet = workbook.add_worksheet('Weekly Schedule')
            
            # Add styling
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#374151',
                'font_color': 'white',
                'border': 1,
                'align': 'center'
            })
            
            cell_format = workbook.add_format({
                'border': 1,
                'align': 'center'
            })
            
            # Write headers
            headers = ['Name', 'Dimanche', 'Lundi', 'Mardi', 'Mercredi', 'Jeudi', 'Vendredi', 'Samedi', 'TOTAL']
            for col, header in enumerate(headers):
                worksheet.write(0, col, header, header_format)
            
            # Set column widths
            worksheet.set_column('A:A', 15)
            worksheet.set_column('B:I', 12)
            
            # Write data
            for row, (index, data) in enumerate(df.iterrows(), 1):
                worksheet.write(row, 0, data.get('Name', ''), cell_format)
                worksheet.write(row, 1, data.get('Sunday', ''), cell_format)
                worksheet.write(row, 2, data.get('Monday', ''), cell_format)
                worksheet.write(row, 3, data.get('Tuesday', ''), cell_format)
                worksheet.write(row, 4, data.get('Wednesday', ''), cell_format)
                worksheet.write(row, 5, data.get('Thursday', ''), cell_format)
                worksheet.write(row, 6, data.get('Friday', ''), cell_format)
                worksheet.write(row, 7, data.get('Saturday', ''), cell_format)
                worksheet.write(row, 8, data.get('TOTAL', ''), cell_format)
            
            workbook.close()
            
            return jsonify({"message": f"Fichier {filename} mis à jour avec succès"})
            
        finally:
            os.chdir(original_cwd)
            
    except Exception as e:
        print(f"Error updating archive file: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/reports/archive/<filename>', methods=['DELETE'])
def delete_archive_file(filename):
    """Delete an archive file"""
    try:
        # Reject temporary Excel files
        if filename.startswith('~$'):
            return jsonify({"error": "Cannot delete temporary Excel files"}), 400
        
        archive_dir = get_archive_directory()
        archive_path = os.path.join(archive_dir, filename)
        
        if os.path.exists(archive_path):
            try:
                os.remove(archive_path)
                return jsonify({"message": f"Fichier {filename} supprimé avec succès"})
            except PermissionError:
                return jsonify({"error": "Impossible de supprimer le fichier. Veuillez fermer le fichier Excel s'il est ouvert."}), 400
            except OSError as e:
                return jsonify({"error": f"Erreur lors de la suppression: {str(e)}"}), 500
        else:
            return jsonify({"error": "Fichier non trouvé"}), 404
            
    except Exception as e:
        print(f"Error deleting archive file: {e}")
        # Provide more user-friendly error messages
        if "Permission denied" in str(e):
            return jsonify({"error": "Impossible de supprimer le fichier. Veuillez fermer le fichier Excel s'il est ouvert."}), 400
        else:
            return jsonify({"error": "Erreur lors de la suppression"}), 500

@app.route('/api/reports/archive/<filename>/download', methods=['GET'])
def download_archive_file(filename):
    """Download an archive file"""
    try:
        from flask import send_file
        
        # Reject temporary Excel files
        if filename.startswith('~$'):
            return jsonify({"error": "Cannot download temporary Excel files"}), 400
        
        archive_dir = get_archive_directory()
        archive_path = os.path.join(archive_dir, filename)
        
        if os.path.exists(archive_path):
            return send_file(archive_path, as_attachment=True, download_name=filename)
        else:
            return jsonify({"error": "Fichier non trouvé"}), 404
            
    except Exception as e:
        print(f"Error downloading archive file: {e}")
        return jsonify({"error": str(e)}), 500

# ==================== TEAMS MANAGEMENT API ====================

@app.route('/api/teams', methods=['GET'])
def get_teams():
    """Get all teams with member and shift counts"""
    try:
        conn = get_db_connection()
        
        # Get teams with member and shift counts
        teams_query = '''
            SELECT t.id, t.name, t.description, t.color, t.created_at,
                   COUNT(DISTINCT tm.employee_id) as member_count,
                   COUNT(DISTINCT ts.id) as shift_count
            FROM teams t
            LEFT JOIN team_members tm ON t.id = tm.team_id
            LEFT JOIN team_shifts ts ON t.id = ts.team_id
            GROUP BY t.id, t.name, t.description, t.color, t.created_at
            ORDER BY t.created_at DESC
        '''
        
        teams = conn.execute(teams_query).fetchall()
        conn.close()
        
        teams_list = []
        for team in teams:
            teams_list.append({
                'id': team['id'],
                'name': team['name'],
                'description': team['description'],
                'color': team['color'],
                'created_at': team['created_at'],
                'member_count': team['member_count'],
                'shift_count': team['shift_count']
            })
        
        return jsonify(teams_list)
        
    except Exception as e:
        print(f"Error fetching teams: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams', methods=['POST'])
def create_team():
    """Create a new team"""
    try:
        data = request.get_json()
        
        if not data or not data.get('name'):
            return jsonify({"error": "Team name is required"}), 400
        
        conn = get_db_connection()
        
        # Check if team name already exists
        existing_team = conn.execute('''
            SELECT id FROM teams WHERE LOWER(name) = LOWER(?)
        ''', (data['name'],)).fetchone()
        
        if existing_team:
            conn.close()
            return jsonify({"error": "A team with this name already exists"}), 400
        
        cursor = conn.execute('''
            INSERT INTO teams (name, description, color)
            VALUES (?, ?, ?)
        ''', (data['name'], data.get('description', ''), data.get('color', '#3b82f6')))
        
        team_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            "message": "Team created successfully",
            "team_id": team_id
        })
        
    except Exception as e:
        print(f"Error creating team: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>', methods=['DELETE'])
def delete_team(team_id):
    """Delete a team and all its associations"""
    try:
        conn = get_db_connection()
        
        # Delete in order: team_shifts, team_members, then team
        conn.execute('DELETE FROM team_shifts WHERE team_id = ?', (team_id,))
        conn.execute('DELETE FROM team_members WHERE team_id = ?', (team_id,))
        conn.execute('DELETE FROM teams WHERE id = ?', (team_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Team deleted successfully"})
        
    except Exception as e:
        print(f"Error deleting team: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/members', methods=['GET'])
def get_team_members(team_id):
    """Get all members of a specific team"""
    try:
        conn = get_db_connection()
        
        members_query = '''
            SELECT tm.id, tm.employee_id, tm.role, w.name, w.employee_number
            FROM team_members tm
            JOIN workers w ON tm.employee_id = w.id
            WHERE tm.team_id = ?
            ORDER BY tm.role, w.name
        '''
        
        members = conn.execute(members_query, (team_id,)).fetchall()
        conn.close()
        
        members_list = []
        for member in members:
            members_list.append({
                'id': member['id'],
                'employee_id': member['employee_id'],
                'name': member['name'],
                'role': member['role'],
                'employee_number': member['employee_number']
            })
        
        return jsonify(members_list)
        
    except Exception as e:
        print(f"Error fetching team members: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/members', methods=['POST'])
def add_team_member(team_id):
    """Add a member to a team"""
    try:
        data = request.get_json()
        
        if not data or not data.get('employee_id'):
            return jsonify({"error": "Employee ID is required"}), 400
        
        conn = get_db_connection()
        
        # Check if employee is already assigned to ANY team
        existing_assignment = conn.execute('''
            SELECT tm.team_id, t.name as team_name
            FROM team_members tm
            JOIN teams t ON tm.team_id = t.id
            WHERE tm.employee_id = ?
        ''', (data['employee_id'],)).fetchone()
        
        if existing_assignment:
            conn.close()
            return jsonify({
                "error": f"L'employé est déjà assigné à l'équipe '{existing_assignment['team_name']}'. Un employé ne peut être assigné qu'à une seule équipe."
            }), 400
        
        # Check if member already exists in current team (double check)
        existing_in_team = conn.execute('''
            SELECT id FROM team_members 
            WHERE team_id = ? AND employee_id = ?
        ''', (team_id, data['employee_id'])).fetchone()
        
        if existing_in_team:
            conn.close()
            return jsonify({"error": "L'employé fait déjà partie de cette équipe"}), 400
        
        # Add member to team
        conn.execute('''
            INSERT INTO team_members (team_id, employee_id, role)
            VALUES (?, ?, ?)
        ''', (team_id, data['employee_id'], data.get('role', 'member')))
        
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Member added successfully"})
        
    except Exception as e:
        print(f"Error adding team member: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/members/<int:member_id>', methods=['DELETE'])
def remove_team_member(team_id, member_id):
    """Remove a member from a team"""
    try:
        conn = get_db_connection()
        
        conn.execute('''
            DELETE FROM team_members 
            WHERE id = ? AND team_id = ?
        ''', (member_id, team_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Member removed successfully"})
        
    except Exception as e:
        print(f"Error removing team member: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/shifts', methods=['GET'])
def get_team_shifts(team_id):
    """Get all shifts for a specific team"""
    try:
        conn = get_db_connection()
        
        shifts_query = '''
            SELECT id, name, start_time, end_time, frequency, start_date, end_date
            FROM team_shifts
            WHERE team_id = ?
            ORDER BY start_time
        '''
        
        shifts = conn.execute(shifts_query, (team_id,)).fetchall()
        conn.close()
        
        shifts_list = []
        for shift in shifts:
            shifts_list.append({
                'id': shift['id'],
                'name': shift['name'],
                'start_time': shift['start_time'],
                'end_time': shift['end_time'],
                'frequency': shift['frequency'],
                'start_date': shift['start_date'],
                'end_date': shift['end_date']
            })
        
        return jsonify(shifts_list)
        
    except Exception as e:
        print(f"Error fetching team shifts: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/shifts', methods=['POST'])
def add_team_shift(team_id):
    """Add a new shift to a team"""
    try:
        data = request.get_json()
        
        required_fields = ['name', 'start_time', 'end_time', 'frequency', 'start_date']
        for field in required_fields:
            if not data or not data.get(field):
                return jsonify({"error": f"{field} is required"}), 400
        
        conn = get_db_connection()
        
        cursor = conn.execute('''
            INSERT INTO team_shifts (team_id, name, start_time, end_time, frequency, start_date, end_date)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (team_id, data['name'], data['start_time'], data['end_time'], 
              data['frequency'], data['start_date'], data.get('end_date')))
        
        shift_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        return jsonify({
            "message": "Shift created successfully",
            "shift_id": shift_id
        })
        
    except Exception as e:
        print(f"Error adding team shift: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/shifts/<int:shift_id>', methods=['DELETE'])
def remove_team_shift(team_id, shift_id):
    """Remove a shift from a team"""
    try:
        conn = get_db_connection()
        
        # Also remove any generated schedules for this shift
        conn.execute('''
            DELETE FROM employee_schedules 
            WHERE id IN (
                SELECT es.id FROM employee_schedules es
                JOIN team_members tm ON es.employee_id = tm.employee_id
                WHERE tm.team_id = ? AND es.schedule_type = 'team_shift'
            )
        ''', (team_id,))
        
        conn.execute('''
            DELETE FROM team_shifts 
            WHERE id = ? AND team_id = ?
        ''', (shift_id, team_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Shift removed successfully"})
        
    except Exception as e:
        print(f"Error removing team shift: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>/shifts/<int:shift_id>/apply', methods=['POST'])
def apply_team_shift(team_id, shift_id):
    """Apply a team shift to all team members"""
    try:
        conn = get_db_connection()
        
        # Get shift details
        shift = conn.execute('''
            SELECT * FROM team_shifts WHERE id = ? AND team_id = ?
        ''', (shift_id, team_id)).fetchone()
        
        if not shift:
            conn.close()
            return jsonify({"error": "Shift not found"}), 404
        
        # Get team members
        members = conn.execute('''
            SELECT employee_id FROM team_members WHERE team_id = ?
        ''', (team_id,)).fetchall()
        
        if not members:
            conn.close()
            return jsonify({"error": "No members in team"}), 400
        
        # Generate schedule entries based on frequency
        from datetime import datetime, timedelta
        
        start_date = datetime.strptime(shift['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(shift['end_date'], '%Y-%m-%d') if shift['end_date'] else start_date + timedelta(days=365)
        
        affected_members = 0
        current_date = start_date
        
        while current_date <= end_date:
            for member in members:
                # Check if schedule already exists for this date
                existing = conn.execute('''
                    SELECT id FROM employee_schedules 
                    WHERE employee_id = ? AND date = ?
                ''', (member['employee_id'], current_date.strftime('%Y-%m-%d'))).fetchone()
                
                if not existing:
                    # Calculate hours worked
                    start_time = datetime.strptime(shift['start_time'], '%H:%M')
                    end_time = datetime.strptime(shift['end_time'], '%H:%M')
                    
                    if end_time < start_time:  # Night shift
                        hours_worked = (24 - start_time.hour + end_time.hour) + (end_time.minute - start_time.minute) / 60
                    else:
                        hours_worked = (end_time.hour - start_time.hour) + (end_time.minute - start_time.minute) / 60
                    
                    # Insert schedule
                    conn.execute('''
                        INSERT INTO employee_schedules 
                        (employee_id, date, day_of_week, start_time, end_time, schedule_type, hours_worked)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (member['employee_id'], current_date.strftime('%Y-%m-%d'), 
                          current_date.weekday(), shift['start_time'], shift['end_time'], 
                          'team_shift', hours_worked))
            
            affected_members = len(members)
            current_date += timedelta(days=int(shift['frequency']))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            "message": "Shift applied successfully",
            "affected_members": affected_members
        })
        
    except Exception as e:
        print(f"Error applying team shift: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/teams/<int:team_id>', methods=['PUT'])
def update_team(team_id):
    """Update a team's details"""
    try:
        data = request.get_json()
        name = data.get('name')
        description = data.get('description', '')
        color = data.get('color', '#3b82f6')
        
        if not name:
            return jsonify({"error": "Team name is required"}), 400
        
        conn = get_db_connection()
        
        # Check if team exists
        existing_team = conn.execute('SELECT id FROM teams WHERE id = ?', (team_id,)).fetchone()
        if not existing_team:
            conn.close()
            return jsonify({"error": "Team not found"}), 404
        
        # Check if another team with the same name exists (excluding current team)
        existing_name = conn.execute('SELECT id FROM teams WHERE name = ? AND id != ?', (name, team_id)).fetchone()
        if existing_name:
            conn.close()
            return jsonify({"error": "A team with this name already exists"}), 400
        
        # Update the team
        conn.execute('''
            UPDATE teams 
            SET name = ?, description = ?, color = ?
            WHERE id = ?
        ''', (name, description, color, team_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({"message": "Team updated successfully"})
        
    except Exception as e:
        print(f"Error updating team: {e}")
        return jsonify({"error": str(e)}), 500

# --- Team Schedule API ---
@app.route('/api/schedule/team-range', methods=['GET'])
def get_team_schedule_range():
    """Get schedule data for all workers within a date range for team schedules"""
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if not start_date_str or not end_date_str:
        return jsonify({"error": "start_date and end_date parameters are required"}), 400
    
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        print(f"Request received for team schedules from {start_date} to {end_date}")
        
        # Get all workers
        conn = get_db_connection()
        workers = conn.execute('SELECT id, name FROM workers ORDER BY name ASC').fetchall()
        
        # Get all schedules for the date range
        schedules = conn.execute('''
            SELECT employee_id, date, start_time, end_time, activity_type, is_night_shift, 
                   hours_worked, schedule_type
            FROM employee_schedules
            WHERE date >= ? AND date <= ?
            ORDER BY employee_id, date
        ''', (start_date_str, end_date_str)).fetchall()
        
        conn.close()
        
        # Build schedule data for each worker
        schedule_data = []
        
        for worker in workers:
            worker_id = worker['id']
            worker_name = worker['name']
            
            # Get schedules for this worker
            worker_schedules = [s for s in schedules if s['employee_id'] == worker_id]
            
            # Build date-keyed schedule data
            worker_schedule_data = {
                'employee_name': worker_name,
                'employee_id': worker_id,
                'schedules': {}
            }
            
            for schedule in worker_schedules:
                schedule_date = schedule['date']
                
                # Handle different schedule types
                if schedule['schedule_type'] == 'work':
                    # Normal work schedule
                    activity = schedule['activity_type'] or 'X'
                    start_time = schedule['start_time']
                    
                    # Normalize activity types to match frontend expectations
                    if activity.upper() == 'X':
                        # Check if this is a late start (after 4 PM)
                        if start_time and start_time >= '16:00':
                            # Late start - show D with hour
                            hour = start_time.split(':')[0]
                            display_value = f"D\n{hour}h"
                        else:
                            display_value = 'X'
                    elif activity.upper() == 'S':
                        display_value = 'S'  
                    elif activity.upper() == 'RP':
                        display_value = 'RP'
                    elif activity.upper() == 'V':
                        display_value = 'V'  # Vacances
                    elif activity.upper() == 'C':
                        display_value = 'C'  # Congé
                    elif activity.upper() == 'M':
                        display_value = 'M'  # Maladie
                    elif activity.upper() == 'D':
                        # D activity type - show with start hour
                        if start_time:
                            hour = start_time.split(':')[0]
                            display_value = f"D\n{hour}h"
                        else:
                            display_value = "D\n17h"  # Default
                    else:
                        # For any other activity type, check if it's a late start
                        if start_time and start_time >= '16:00':
                            hour = start_time.split(':')[0]
                            display_value = f"D\n{hour}h"
                        else:
                            display_value = activity.upper()
                
                elif schedule['schedule_type'] == 'unavailable':
                    # Handle unavailable schedules (vacation, sick leave, etc.)
                    activity = schedule['activity_type'] or 'V'
                    if activity.upper() == 'V':
                        display_value = 'V'  # Vacances
                    elif activity.upper() == 'C':
                        display_value = 'C'  # Congé
                    elif activity.upper() == 'M':
                        display_value = 'M'  # Maladie
                    else:
                        display_value = activity.upper()
                
                elif schedule['schedule_type'] == 'end_of_night_shift':
                    # Night shift ending - show F + time only if end time is before 12:00 PM (noon)
                    # Otherwise, show as regular X shift
                    end_time = schedule['end_time']
                    if end_time:
                        # Check if end time is before 12:00 PM (noon)
                        hour = int(end_time.split(':')[0])
                        if hour < 12:  # Before noon, use F\n format
                            display_value = f"F\n{hour}h"
                        else:  # At or after noon, use X format
                            display_value = "X"
                    else:
                        display_value = "F\n9h"  # Default
                else:
                    continue  # Skip unknown schedule types
                
                # Store the schedule data
                worker_schedule_data['schedules'][schedule_date] = {
                    'activity_type': display_value,
                    'start_time': schedule['start_time'],
                    'end_time': schedule['end_time'],
                    'hours_worked': schedule['hours_worked'],
                    'is_night_shift': schedule['is_night_shift'],
                    'raw_text': f"{display_value} {schedule['start_time']}-{schedule['end_time']}" if schedule['start_time'] and schedule['end_time'] else display_value
                }
            
            schedule_data.append(worker_schedule_data)
        
        print(f"Generated team schedule data for {len(schedule_data)} workers covering {len([d for d in schedule_data if d['schedules']])} workers with schedules")
        return jsonify(schedule_data)
            
    except Exception as e:
        print(f"Error generating team schedule range: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

# --- Worker Schedule API ---
@app.route('/api/workers/<worker_identifier>/schedule/day', methods=['GET'])
def get_worker_schedule_by_date(worker_identifier):
    """Get a worker's schedule for a specific date"""
    date_str = request.args.get('date')
    if not date_str:
        return jsonify({"error": "Date parameter is required"}), 400
    
    try:
        # Convert date string to proper format
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        conn = get_db_connection()
        
        # First try to find worker by ID (if numeric)
        worker_id = None
        try:
            worker_id = int(worker_identifier)
            worker = conn.execute('SELECT * FROM workers WHERE id = ?', (worker_id,)).fetchone()
        except ValueError:
            # If not numeric, find by name
            worker = conn.execute('SELECT * FROM workers WHERE LOWER(name) = LOWER(?)', (worker_identifier,)).fetchone()
            if worker:
                worker_id = worker['id']
        
        if not worker_id:
            return jsonify({"error": "Worker not found"}), 404
        
        # Get schedule for the specific date
        schedule = conn.execute('''
            SELECT employee_id, date, start_time, end_time, activity_type, 
                   is_night_shift, hours_worked, schedule_type
            FROM employee_schedules 
            WHERE employee_id = ? AND date = ?
        ''', (worker_id, date_str)).fetchall()
        
        conn.close()
        
        # Convert to JSON-serializable format
        schedule_data = []
        for row in schedule:
            schedule_data.append({
                'employee_id': row['employee_id'],
                'date': row['date'],
                'start_time': row['start_time'],
                'end_time': row['end_time'],
                'activity_type': row['activity_type'],
                'is_night_shift': row['is_night_shift'],
                'hours_worked': row['hours_worked'],
                'schedule_type': row['schedule_type']
            })
        
        return jsonify(schedule_data)
        
    except Exception as e:
        print(f"Error getting worker schedule: {e}")
        return jsonify({"error": str(e)}), 500
    

from flask import send_from_directory

# Get absolute path to frontend directory
def get_frontend_path():
    """Get frontend path that works in both development and packaged mode"""
    if getattr(sys, '_MEIPASS', None):
        # Running from PyInstaller bundle - frontend should be in _MEIPASS
        return os.path.join(sys._MEIPASS, 'frontend')
    else:
        # Running from source - use absolute path
        return os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'frontend'))

FRONTEND_DIR = get_frontend_path()
print(f"FRONTEND_DIR resolved to: {FRONTEND_DIR}")
print(f"Frontend index.html exists: {os.path.exists(os.path.join(FRONTEND_DIR, 'index.html'))}")

@app.route('/')
def serve_index():
    try:
        print(f"Serving index from: {FRONTEND_DIR}")
        return send_from_directory(FRONTEND_DIR, 'index.html')
    except Exception as e:
        print(f"Error serving index: {e}")
        return f"Error: {e}", 500

@app.route('/<path:path>')
def serve_static(path):
    try:
        print(f"Serving static file: {path} from: {FRONTEND_DIR}")
        return send_from_directory(FRONTEND_DIR, path)
    except Exception as e:
        print(f"Error serving {path}: {e}")
        return f"Error: {e}", 500



# --- Run the Application ---
if __name__ == '__main__':
    # Ensure the database schema is up-to-date before running
    migrate_database()
    
    # Run the Flask app
    # Using a non-standard port to avoid conflicts
    app.run(debug=True, port=5001)